"""
Dependency Analyzer - Analyzes C++ file dependencies and function relationships
"""

import os
import re
from typing import List, Dict, Set


class DependencyAnalyzer:
    """Analyzes C++ file dependencies and function relationships"""

    def __init__(self, repo_path: str, base_path: str = None):
        self.repo_path = os.path.abspath(repo_path)
        self.base_path = os.path.abspath(base_path) if base_path else self.repo_path
        # Configurable hardware denylist
        self.hardware_symbols = {
            "digitalWrite", "digitalRead", "delay", "millis",
            "Serial.", "WiFi.", "SPIFFS", "HTTPClient",
            "pinMode", "analogRead", "analogWrite", "tone", "noTone",
            "attachInterrupt", "detachInterrupt", "interrupts", "noInterrupts",
            "pulseIn", "pulseInLong", "shiftIn", "shiftOut",
            "SPI.", "Wire.", "I2C", "CAN", "RTOS", "FreeRTOS",
            "xTaskCreate", "vTaskDelay", "vTaskDelete"
        }
        self.hardware_headers = {
            "Arduino.h", "WiFi.h", "SPIFFS.h", "HTTPClient.h",
            "SPI.h", "Wire.h", "CAN.h", "freertos/FreeRTOS.h"
        }

    def find_all_c_files(self) -> List[str]:
        """Find all C/C++ files in the repository"""
        c_files = []
        for root, dirs, files in os.walk(self.repo_path):
            for file in files:
                if file.endswith(('.c', '.cpp', '.h', '.hpp')):
                    # Skip files in tests/, build/, CMakeFiles/, and ai_test_build/ directories
                    if any(skip_dir in root.split(os.sep) for skip_dir in ['tests', 'build', 'CMakeFiles', 'ai_test_build']):
                        continue
                    # Skip main.cpp as it's not suitable for unit testing
                    if file == 'main.cpp':
                        continue
                    c_files.append(os.path.join(root, file))
        return c_files

    def perform_repo_scan(self) -> Dict:
        """Perform lightweight static repo scan for dependency analysis"""
        print("[INFO] [REPO SCAN] Starting lightweight repo scan...")
        
        # Phase 1: File Index
        file_index = self._build_file_index()
        print(f"[INFO] [REPO SCAN] Indexed {len(file_index)} files: {', '.join(file_index.keys())}")
        
        # Phase 2: Function Index
        function_index = self._build_function_index(file_index)
        print(f"[INFO] [REPO SCAN] Indexed {len(function_index)} functions:")
        for func_name, func_info in function_index.items():
            print(f"  - {func_name} (from {func_info['file']})")
        
        # Phase 2.5: Class Role Classification
        class_roles = self._classify_classes(function_index, file_index)
        print(f"[INFO] [REPO SCAN] Classified {len(class_roles)} classes:")
        for cls, role in class_roles.items():
            print(f"  - {cls}: {role}")
        
        # Phase 3: Hardware Detection
        hardware_flags = self._detect_hardware(function_index, class_roles, file_index)
        hw_funcs = [f for f, hw in hardware_flags.items() if hw]
        print(f"[INFO] [REPO SCAN] Marked {len(hw_funcs)} functions as hardware-touching:")
        for func_name in hw_funcs:
            func_info = function_index[func_name]
            print(f"  - {func_name} (from {func_info['file']})")
        print("[INFO] [REPO SCAN] Note: Hardware classification based on class roles, not API usage.")
        
        # Phase 4: Call Depth Computation
        call_graph = self._build_call_graph(function_index)
        call_depths = self._compute_call_depths(call_graph, hardware_flags)
        print(f"[INFO] [REPO SCAN] Computed call depths (max depth: {max(call_depths.values()) if call_depths else 0})")
        depth_examples = [(f, d) for f, d in call_depths.items() if d > 0][:5]  # Show first 5 with depth > 0
        if depth_examples:
            print("  Sample call depths:")
            for func, depth in depth_examples:
                print(f"    - {func}: depth {depth}")
        
        # Phase 5: File-Level Summary
        file_summaries = self._build_file_summaries(file_index, function_index, hardware_flags, call_depths)
        print(f"[INFO] [REPO SCAN] Built summaries for {len(file_summaries)} files:")
        for file_path, summary in file_summaries.items():
            testable = summary.get('testable', False)
            hw_free = summary.get('hardware_free', False)
            max_depth = summary.get('max_call_depth', 0)
            func_count = len(summary.get('functions', []))
            print(f"  - {file_path}: {func_count} functions, testable={testable}, hardware_free={hw_free}, max_depth={max_depth}")
        
        return {
            'file_index': file_index,
            'function_index': function_index,
            'class_roles': class_roles,
            'call_graph': call_graph,
            'hardware_flags': hardware_flags,
            'call_depths': call_depths,
            'file_summaries': file_summaries
        }

    def save_repo_scan_results(self, scan_results: Dict, output_dir: str = None):
        """Save detailed repo scan results to text files"""
        import os
        if output_dir is None:
            # Find repo root (parent of repo_path if repo_path is a subdirectory)
            repo_root = self.repo_path
            if os.path.exists(os.path.join(self.repo_path, '.git')):
                repo_root = self.repo_path
            else:
                # Try to find git root by going up directories
                current = self.repo_path
                for _ in range(3):  # Go up max 3 levels
                    parent = os.path.dirname(current)
                    if os.path.exists(os.path.join(parent, '.git')):
                        repo_root = parent
                        break
                    current = parent
            
            output_dir = os.path.join(repo_root, "test_analysis")
        os.makedirs(output_dir, exist_ok=True)
        
        # Save function index
        with open(os.path.join(output_dir, "functions.txt"), "w") as f:
            f.write("REPO SCAN - FUNCTION INDEX\n")
            f.write("=" * 50 + "\n\n")
            for func_name, func_info in scan_results['function_index'].items():
                f.write(f"Function: {func_name}\n")
                f.write(f"File: {func_info['file']}\n")
                f.write(f"Language: {func_info['language']}\n")
                f.write(f"Calls: {', '.join(func_info['calls'])}\n")
                f.write(f"Hardware Touching: {scan_results['hardware_flags'].get(func_name, False)}\n")
                f.write(f"Call Depth: {scan_results['call_depths'].get(func_name, 0)}\n")
                f.write("-" * 30 + "\n\n")
        
        # Save hardware functions
        with open(os.path.join(output_dir, "hardware_functions.txt"), "w") as f:
            f.write("REPO SCAN - HARDWARE-TOUCHING FUNCTIONS\n")
            f.write("=" * 50 + "\n\n")
            hw_funcs = [f for f, hw in scan_results['hardware_flags'].items() if hw]
            for func_name in hw_funcs:
                func_info = scan_results['function_index'][func_name]
                f.write(f"{func_name} (from {func_info['file']})\n")
        
        # Save file summaries
        with open(os.path.join(output_dir, "file_summaries.txt"), "w") as f:
            f.write("REPO SCAN - FILE SUMMARIES\n")
            f.write("=" * 50 + "\n\n")
            for file_path, summary in scan_results['file_summaries'].items():
                f.write(f"File: {file_path}\n")
                f.write(f"Functions: {len(summary.get('functions', []))}\n")
                f.write(f"Testable: {summary.get('testable', False)}\n")
                f.write(f"Hardware Free: {summary.get('hardware_free', False)}\n")
                f.write(f"Max Call Depth: {summary.get('max_call_depth', 0)}\n")
                f.write(f"Dependencies: {', '.join(summary.get('dependencies', []))}\n")
                f.write("-" * 30 + "\n\n")
        
        # Save call depths
        with open(os.path.join(output_dir, "call_depths.txt"), "w") as f:
            f.write("REPO SCAN - CALL DEPTHS\n")
            f.write("=" * 50 + "\n\n")
            for func_name, depth in sorted(scan_results['call_depths'].items(), key=lambda x: x[1], reverse=True):
                if depth > 0:
                    func_info = scan_results['function_index'][func_name]
                    f.write(f"Depth {depth}: {func_name} (from {func_info['file']})\n")
        
        print(f"[INFO] [REPO SCAN] Results saved to {output_dir}/")

    def export_to_excel(self, scan_results: Dict, output_file: str):
        """Export scan results to Excel format for better visualization"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            print("[ERROR] openpyxl not installed. Install with: pip install openpyxl")
            return

        wb = Workbook()
        
        # Sheet 1: Function Index
        ws1 = wb.active
        ws1.title = "Function Index"
        ws1['A1'] = "Function Name"
        ws1['B1'] = "File"
        ws1['C1'] = "Language"
        ws1['D1'] = "Calls"
        ws1['E1'] = "Hardware Touching"
        ws1['F1'] = "Call Depth"
        
        # Style header
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        for cell in ws1[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        row = 2
        for func_name, func_info in scan_results['function_index'].items():
            ws1[f'A{row}'] = func_name
            ws1[f'B{row}'] = func_info['file']
            ws1[f'C{row}'] = func_info['language']
            ws1[f'D{row}'] = ', '.join(func_info['calls'])
            ws1[f'E{row}'] = scan_results['hardware_flags'].get(func_name, False)
            ws1[f'F{row}'] = scan_results['call_depths'].get(func_name, 0)
            row += 1
        
        # Auto-adjust column widths
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws1.column_dimensions[col].width = 20
        
        # Sheet 2: Call Graph
        ws2 = wb.create_sheet("Call Graph")
        ws2['A1'] = "Caller Function"
        ws2['B1'] = "Callee Functions"
        for cell in ws2[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        row = 2
        for caller, callees in scan_results['call_graph'].items():
            ws2[f'A{row}'] = caller
            ws2[f'B{row}'] = ', '.join(callees)
            row += 1
        
        ws2.column_dimensions['A'].width = 30
        ws2.column_dimensions['B'].width = 50
        
        # Sheet 3: File Summaries
        ws3 = wb.create_sheet("File Summaries")
        ws3['A1'] = "File Path"
        ws3['B1'] = "Function Count"
        ws3['C1'] = "Testable"
        ws3['D1'] = "Hardware Free"
        ws3['E1'] = "Max Call Depth"
        ws3['F1'] = "Dependencies"
        for cell in ws3[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        row = 2
        for file_path, summary in scan_results['file_summaries'].items():
            ws3[f'A{row}'] = file_path
            ws3[f'B{row}'] = len(summary.get('functions', []))
            ws3[f'C{row}'] = summary.get('testable', False)
            ws3[f'D{row}'] = summary.get('hardware_free', False)
            ws3[f'E{row}'] = summary.get('max_call_depth', 0)
            ws3[f'F{row}'] = ', '.join(summary.get('dependencies', []))
            row += 1
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws3.column_dimensions[col].width = 20
        
        # Sheet 4: Class Roles
        ws4 = wb.create_sheet("Class Roles")
        ws4['A1'] = "Class Name"
        ws4['B1'] = "Role"
        for cell in ws4[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        row = 2
        for class_name, role in scan_results['class_roles'].items():
            ws4[f'A{row}'] = class_name
            ws4[f'B{row}'] = role
            row += 1
        
        ws4.column_dimensions['A'].width = 25
        ws4.column_dimensions['B'].width = 15
        
        # Save the workbook
        wb.save(output_file)
        print(f"[INFO] Excel export saved to {output_file}")

    def _build_file_index(self) -> Dict[str, Dict]:
        """Phase 1: Build file index"""
        file_index = {}
        all_files = self.find_all_c_files()
        
        for file_path in all_files:
            rel_path = os.path.relpath(file_path, self.base_path)
            ext = os.path.splitext(file_path)[1].lower()
            language = 'c' if ext == '.c' else 'cpp'
            is_header = ext in ['.h', '.hpp']
            
            includes = self._extract_includes(file_path)
            
            file_index[rel_path] = {
                'path': rel_path,
                'language': language,
                'includes': includes,
                'is_header': is_header
            }
        
        return file_index

    def _build_function_index(self, file_index: Dict) -> Dict[str, Dict]:
        """Phase 2: Build function index"""
        function_index = {}
        
        for file_path, file_info in file_index.items():
            if file_info['is_header']:
                continue  # Skip headers for function definitions
            
            abs_path = os.path.join(self.base_path, file_path)
            functions = self._extract_functions(abs_path)
            
            for func in functions:
                func_name = func['name']
                function_index[func_name] = {
                    'name': func_name,
                    'file': file_path,
                    'language': file_info['language'],
                    'calls': func['calls'],
                    'body': func['body'],
                    'touches_hardware': False  # Will be set in Phase 3
                }
        
        return function_index

    def _classify_classes(self, function_index: Dict, file_index: Dict) -> Dict[str, str]:
        """Phase 2.5: Classify classes by role (HARDWARE, LOGIC, MIXED)"""
        classes = set()
        for func_name in function_index:
            if '::' in func_name:
                # Extract class name from "ReturnType ClassName::method"
                parts = func_name.split('::')[0].split()
                if parts:
                    class_name = parts[-1]
                    classes.add(class_name)
        
        class_roles = {}
        for cls in classes:
            cls_lower = cls.lower()
            # Treat interface-style abstractions as testable boundaries.
            # Example: IGpio should not make callers look "hardware-only".
            is_interface_like = bool(re.match(r"^I[A-Z][a-z]", cls))

            if not is_interface_like and any(keyword in cls_lower for keyword in ['hardware', 'hal', 'driver', 'io', 'gpio']):
                class_roles[cls] = 'HARDWARE'
            elif any(keyword in cls_lower for keyword in ['parser', 'validator']):
                class_roles[cls] = 'LOGIC'
            else:
                class_roles[cls] = 'MIXED'  # Default for managers, state machines, etc.
        
        # Override based on file names
        for file_path in file_index:
            file_lower = file_path.lower()
            if 'hardware' in file_lower or 'hal' in file_lower:
                for cls in classes:
                    if cls.lower() in file_lower:
                        class_roles[cls] = 'HARDWARE'
        
        return class_roles

    def _detect_hardware(self, function_index: Dict, class_roles: Dict, file_index: Dict) -> Dict[str, bool]:
        """Phase 3: Detect hardware-touching functions based on class roles, code analysis, and headers"""
        hardware_flags = {}
        
        for func_name, func_info in function_index.items():
            is_hardware = False
            
            # Check 1: Class role-based detection
            class_name = None
            if '::' in func_name:
                parts = func_name.split('::')[0].split()
                if parts:
                    class_name = parts[-1]
            
            if class_name and class_name in class_roles:
                role = class_roles[class_name]
                if role == 'HARDWARE':
                    is_hardware = True
                elif role == 'LOGIC':
                    is_hardware = False
                # For MIXED, continue with additional checks below
            
            # Check 2: Direct hardware symbol detection in function body
            if not is_hardware and 'body' in func_info:
                body = func_info['body']
                for symbol in self.hardware_symbols:
                    if symbol in body:
                        is_hardware = True
                        break
            
            # Check 3: Hardware header detection
            if not is_hardware:
                file_path = func_info['file']
                if file_path in file_index:
                    includes = file_index[file_path]['includes']
                    for header in self.hardware_headers:
                        if any(header in inc for inc in includes):
                            is_hardware = True
                            break
            
            # Check 4: Qualified call analysis (conservative)
            # The call list is untyped and can match unrelated methods (e.g., a DI call like gpio.write()
            # matching some unrelated Gpio::write()). Only treat it as hardware when the function body
            # contains an explicit qualified call like HwClass::method.
            if not is_hardware and class_name and class_name in class_roles and class_roles[class_name] == 'MIXED':
                body = func_info.get('body', '') or ''
                for call in func_info.get('calls', []):
                    for func_key in function_index:
                        if not func_key.endswith('::' + call):
                            continue
                        call_class = func_key.split('::')[0].split()[-1]
                        if class_roles.get(call_class) != 'HARDWARE':
                            continue
                        if re.search(rf"\\b{re.escape(call_class)}\\s*::\\s*{re.escape(call)}\\b", body):
                            is_hardware = True
                            break
                    if is_hardware:
                        break
            
            hardware_flags[func_name] = is_hardware
        
        return hardware_flags
        
        # First pass: Mark functions that directly touch hardware
        for func_name, func_info in function_index.items():
            # Check function body for hardware symbols
            if 'body' in func_info:
                for symbol in self.hardware_symbols:
                    if symbol in func_info['body']:
                        hardware_deps[func_name].add(symbol)
            
            # Check for hardware headers
            file_path = func_info['file']
            if file_path in file_index:
                includes = file_index[file_path]['includes']
                for header in self.hardware_headers:
                    if any(header in inc for inc in includes):
                        hardware_deps[func_name].add(header)
        
        # Propagate upward in call graph
        changed = True
        while changed:
            changed = False
            for func_name, func_info in function_index.items():
                current_deps = hardware_deps[func_name].copy()
                for call in func_info['calls']:
                    if call in hardware_deps:
                        callee_deps = hardware_deps[call]
                        if not callee_deps.issubset(current_deps):
                            hardware_deps[func_name].update(callee_deps)
                            current_deps.update(callee_deps)
                            changed = True
        
        return hardware_deps

    def get_file_analysis(self, file_path: str, scan_results: Dict = None) -> Dict:
        """Generate detailed analysis for a single file in the requested JSON format"""
        rel_path = os.path.normpath(os.path.relpath(file_path, self.repo_path))

        # Ensure we have a full scan first (or at least for this file and its deps)
        if scan_results is None:
            scan_results = self.perform_repo_scan()

        file_functions = [
            f
            for f, info in scan_results['function_index'].items()
            if os.path.normpath(info.get('file', '')) == rel_path
        ]
        
        # Define supported stubs (these are considered testable hardware dependencies)
        supported_stubs = {
            "digitalWrite", "digitalRead", "delay", "millis",
            "Serial.", "pinMode", "analogRead", "analogWrite",
            "SPIFFS", "HTTPClient", "SPIFFS.h", "HTTPClient.h", "Arduino.h",
            "WiFi.", "WiFi.h" # Assuming we might stub WiFi too, but sticking to confirmed ones
        }
        
        functions_data = []
        for func_name in file_functions:
            func_info = scan_results['function_index'].get(func_name, {})
            file_info = scan_results.get('file_index', {}).get(func_info.get('file', ''), {})

            hw_deps_set = set()
            body = func_info.get('body', '') or ''
            for symbol in self.hardware_symbols:
                if symbol in body:
                    hw_deps_set.add(symbol)

            includes = file_info.get('includes', []) or []
            for header in self.hardware_headers:
                if any(header in inc for inc in includes):
                    hw_deps_set.add(header)

            is_hw_flagged = bool(scan_results.get('hardware_flags', {}).get(func_name, False))
            hw_deps = sorted(hw_deps_set)
            call_depth = scan_results['call_depths'].get(func_name, 0)
            
            # Classify function category
            if hw_deps or is_hw_flagged:
                category = "hardware"
                is_testable = False
                reason = f"Hardware-dependent: {', '.join(hw_deps) if hw_deps else 'hardware classification'}"
            else:
                category = "software-only"
                is_testable = True
                reason = "Pure software logic"
            
            functions_data.append({
                "name": func_name,
                "category": category,
                "hardware_calls": hw_deps,
                "call_depth": call_depth,
                "testable": is_testable,
                "reason": reason
            })
            
        return {
            "file": os.path.basename(file_path),
            "functions": functions_data
        }


    def _build_call_graph(self, function_index: Dict) -> Dict[str, List[str]]:
        """Build compact call graph (adjacency list)"""
        call_graph = {}
        for func_name, func_info in function_index.items():
            call_graph[func_name] = func_info['calls']
        return call_graph
    
    def _compute_call_depths(self, call_graph: Dict, hardware_flags: Dict) -> Dict[str, int]:
        """Phase 4: Compute bounded call depths"""
        call_depths = {}
        
        def get_depth(func_name: str, visited: Set[str]) -> int:
            if func_name in visited:
                return 0  # Avoid cycles
            if func_name not in call_graph:
                return 0
            
            visited.add(func_name)
            max_child_depth = 0
            for callee in call_graph[func_name]:
                if not hardware_flags.get(callee, False):  # Only count non-hardware calls
                    child_depth = get_depth(callee, visited.copy())
                    max_child_depth = max(max_child_depth, child_depth)
            
            return 1 + max_child_depth
        
        for func_name in call_graph:
            call_depths[func_name] = get_depth(func_name, set())
        
        return call_depths

    def _build_file_summaries(self, file_index: Dict, function_index: Dict, hardware_flags: Dict, call_depths: Dict) -> Dict[str, Dict]:
        """Phase 5: Build file-level summaries"""
        file_summaries = {}
        
        for file_path, file_info in file_index.items():
            if file_info['is_header']:
                continue
            
            functions_in_file = [f for f, info in function_index.items() if info['file'] == file_path]
            max_depth = max((call_depths.get(f, 0) for f in functions_in_file), default=0)
            hardware_free = all(not hardware_flags.get(f, False) for f in functions_in_file)
            
            # Dependencies: files containing called functions within depth 2
            dependencies = set()
            for func in functions_in_file:
                if call_depths.get(func, 0) <= 2:
                    for callee in function_index[func]['calls']:
                        if callee in function_index:
                            dep_file = function_index[callee]['file']
                            if dep_file != file_path:
                                dependencies.add(dep_file)
            
            testable = file_info['language'] == 'cpp' or (file_info['language'] == 'c' and hardware_free)  # Default: exclude C files unless pure logic
            
            file_summaries[file_path] = {
                'testable': testable,
                'functions': functions_in_file,
                'max_call_depth': max_depth,
                'hardware_free': hardware_free,
                'dependencies': list(dependencies)
            }
        
        return file_summaries

    def _extract_includes(self, file_path: str) -> List[str]:
        """Extract include directives from a file"""
        includes = []
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                for line in f:
                    line = line.strip()
                    if line.startswith('#include'):
                        # Extract the include path
                        match = re.search(r'#include\s*[<"]([^>"]+)[>"]', line)
                        if match:
                            includes.append(match.group(1))
        except Exception as e:
            print(f"[WARNING] Could not read {file_path}: {e}")
        return includes

    def _extract_functions(self, file_path: str) -> List[Dict]:
        """Extract function definitions from a file"""
        functions = []
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
            
            # Simple regex-based function extraction (not perfect but sufficient for static analysis)
            # Match function definitions: return_type function_name(parameters) or class::method
            func_pattern = r'(?:^|\n)\s*([a-zA-Z_][a-zA-Z0-9_*\s]+(?:\s*::\s*[a-zA-Z_][a-zA-Z0-9_]*)*)\s*\(([^)]*)\)\s*\{'
            matches = re.finditer(func_pattern, content, re.MULTILINE)
            
            for match in matches:
                func_signature = match.group(1).strip()
                params = match.group(2).strip()
                
                # Parse the signature
                if '::' in func_signature:
                    # Class method: Class::method
                    parts = func_signature.split('::')
                    if len(parts) == 2:
                        class_name = parts[0].strip()
                        method_name = parts[1].strip()
                        return_type = "void"  # Default assumption
                        func_name = f"{class_name}::{method_name}"
                    else:
                        continue
                else:
                    # Plain function: return_type function_name
                    parts = func_signature.split()
                    if len(parts) >= 2:
                        return_type = ' '.join(parts[:-1])
                        func_name = parts[-1]
                    else:
                        continue
                
                # Skip constructors, destructors, operators
                if func_name.startswith('~') or 'operator' in func_name or func_name.split('::')[-1] == class_name.split()[-1] if '::' in func_name else False:
                    continue
                
                # Extract called functions (simple heuristic)
                func_body_start = match.end()
                brace_count = 1
                func_body = ""
                i = func_body_start
                while i < len(content) and brace_count > 0:
                    if content[i] == '{':
                        brace_count += 1
                    elif content[i] == '}':
                        brace_count -= 1
                    if brace_count > 0:
                        func_body += content[i]
                    i += 1
                
                # Extract function calls (very simple)
                calls = []
                call_matches = re.findall(r'\b([a-zA-Z_][a-zA-Z0-9_]*)\s*\(', func_body)
                for call in call_matches:
                    if call not in ['if', 'while', 'for', 'switch', 'return', 'sizeof', 'malloc', 'free'] and len(call) > 1:
                        calls.append(call)
                
                functions.append({
                    'name': func_name,
                    'return_type': return_type,
                    'parameters': params,
                    'calls': list(set(calls)),  # Remove duplicates
                    'body': func_body
                })
                
        except Exception as e:
            print(f"[WARNING] Could not parse {file_path}: {e}")
        
        return functions