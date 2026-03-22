import argparse
import json
import os
import sys
from pathlib import Path
from typing import Any

# Support running as an installed module (preferred) and as a direct script.
try:
    from .analyzer import DependencyAnalyzer
    from .mcdc import analyze_repo_mcdc
except ImportError:
    # Direct execution (e.g. `python CW_Test_Analyzer/ai_c_test_analyzer/cli.py ...`)
    # has no package context, so relative imports fail.
    this_file = Path(__file__).resolve()
    pkg_parent = this_file.parents[1]  # .../CW_Test_Analyzer
    if str(pkg_parent) not in sys.path:
        sys.path.insert(0, str(pkg_parent))
    from ai_c_test_analyzer.analyzer import DependencyAnalyzer
    from ai_c_test_analyzer.mcdc import analyze_repo_mcdc

from .safety_policy import SafetyPolicy, save_safety_summary


def _export_mcdc_gaps_to_excel(payload: dict[str, Any], excel_path: Path) -> None:
    # Lazy import so analyzer still works even if Excel deps are missing.
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    bold = Font(bold=True)

    version = payload.get('version')
    generated_at = payload.get('generated_at')
    source_root = payload.get('source_root')
    files = payload.get('files') or {}

    total_files = 0
    total_decisions = 0
    if isinstance(files, dict):
        total_files = len(files)
        for _, decisions in files.items():
            if isinstance(decisions, list):
                total_decisions += len(decisions)

    rows = [
        ("version", version),
        ("generated_at", generated_at),
        ("source_root", source_root),
        ("files_with_gaps", total_files),
        ("decisions_with_gaps", total_decisions),
    ]

    for r_idx, (k, v) in enumerate(rows, start=1):
        ws_summary.cell(row=r_idx, column=1, value=k).font = bold
        ws_summary.cell(row=r_idx, column=2, value=v)

    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 90
    ws_summary['B3'].alignment = Alignment(wrap_text=True)

    # Detailed gaps sheet
    ws = wb.create_sheet(title="Gaps")
    headers = [
        "file",
        "kind",
        "line",
        "expression",
        "conditions",
        "required_pairs_estimate",
    ]
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.font = bold

    ws.freeze_panes = "A2"

    row = 2
    if isinstance(files, dict):
        for file_path, decisions in sorted(files.items(), key=lambda kv: str(kv[0])):
            if not isinstance(decisions, list):
                continue
            for d in decisions:
                if not isinstance(d, dict):
                    continue
                conditions = d.get('conditions')
                if isinstance(conditions, list):
                    conditions_str = " | ".join(str(x) for x in conditions)
                else:
                    conditions_str = str(conditions) if conditions is not None else ""

                ws.cell(row=row, column=1, value=str(file_path))
                ws.cell(row=row, column=2, value=d.get('kind'))
                ws.cell(row=row, column=3, value=d.get('line'))
                ws.cell(row=row, column=4, value=d.get('expression'))
                ws.cell(row=row, column=5, value=conditions_str)
                ws.cell(row=row, column=6, value=d.get('required_pairs_estimate'))
                row += 1

    ws.column_dimensions['A'].width = 55
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 60
    ws.column_dimensions['E'].width = 60
    ws.column_dimensions['F'].width = 22
    for r in range(2, row):
        ws.cell(row=r, column=4).alignment = Alignment(wrap_text=True)
        ws.cell(row=r, column=5).alignment = Alignment(wrap_text=True)

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(excel_path)

def validate_environment():
    """Validate that required tools and dependencies are available."""
    # Check for Python version
    if sys.version_info < (3, 8):
        print("Error: Python 3.8 or higher is required.")
        sys.exit(1)

def main():
    parser = argparse.ArgumentParser(description="AI C Test Analyzer")
    parser.add_argument('--repo-path', required=True, help='Path to the repository')
    parser.add_argument('--verbose', action='store_true', help='Enable verbose output')
    parser.add_argument('--wait-before-exit', action='store_true', help='Wait for user input before exiting')
    parser.add_argument('--no-excel-output', action='store_true', help='Disable Excel output')
    parser.add_argument(
        '--text-output',
        action='store_true',
        help='Enable auxiliary text outputs (functions.txt, file_summaries.txt, etc.)',
    )
    parser.add_argument('--mcdc', action='store_true', help='Generate MC/DC gap analysis report (tests/analysis/mcdc_gaps.json)')
    parser.add_argument(
        '--safety-level',
        choices=list(SafetyPolicy.allowed_levels()),
        default='QM',
        help=(
            'Configures which analyses, test types, and review gates are required so generated tests align with SIL expectations '
            'without claiming certification.'
        ),
    )
    parser.add_argument('--policy-file', default=None)
    parser.add_argument('--disable-mcdc', action='store_true')
    
    args = parser.parse_args()
    
    if args.verbose:
        print(f"Repository path: {args.repo_path}")
    
    repo_path = Path(args.repo_path)
    if not repo_path.exists():
        print(f"Error: Repository path '{repo_path}' does not exist.")
        sys.exit(1)
    
    policy = SafetyPolicy.load(
        safety_level=args.safety_level,
        repo_root=repo_path,
        policy_file=args.policy_file,
        disable_mcdc=bool(args.disable_mcdc),
    )

    # Enforce mandatory analysis per safety level.
    want_mcdc = bool(args.mcdc) or policy.mcdc_analysis_required()
    if args.disable_mcdc:
        want_mcdc = False

    analyze_repo(
        repo_path,
        args.verbose,
        args.no_excel_output,
        mcdc=want_mcdc,
        text_output=bool(args.text_output),
    )

    # Best-effort: update safety summary.
    try:
        update: dict[str, object] = {
            'safety_level': policy.safety_level,
            'mcdc_analysis_performed': bool(want_mcdc),
        }
        if want_mcdc:
            gaps_path = repo_path / 'tests' / 'analysis' / 'mcdc_gaps.json'
            gaps_remaining = 0
            if gaps_path.exists():
                try:
                    gaps = json.loads(gaps_path.read_text(encoding='utf-8')) or {}
                    for _, decisions in (gaps.get('files', {}) or {}).items():
                        if isinstance(decisions, list):
                            gaps_remaining += len(decisions)
                except Exception:
                    gaps_remaining = 0
            update['mcdc_gaps_remaining'] = gaps_remaining
            update['coverage_status'] = {'mcdc': 'INCOMPLETE' if gaps_remaining else 'PASS'}

        save_safety_summary(repo_path, update)
    except Exception:
        pass
    
    if args.wait_before_exit:
        input("Press Enter to exit...")

def analyze_repo(
    repo_path,
    verbose: bool = False,
    no_excel_output: bool = False,
    *,
    mcdc: bool = False,
    text_output: bool = False,
):
    """Analyze all C/C++ files in the repository."""
    analyzer = DependencyAnalyzer(str(repo_path))
    scan_results = analyzer.perform_repo_scan()
    
    output_dir = repo_path / 'tests' / 'analysis'
    output_dir.mkdir(parents=True, exist_ok=True)

    # Also write a canonical copy under the workspace 'work/' folder so other
    # tools (orchestrator, generator) have one authoritative artifact location.
    # Use the provided repo_path to determine the workspace root (repo parent).
    try:
        workspace_root = Path(repo_path).resolve().parent
    except Exception:
        workspace_root = Path(__file__).resolve().parents[2]
    work_dir = workspace_root / 'work'
    work_dir.mkdir(parents=True, exist_ok=True)
    
    # Optional auxiliary text outputs (often duplicate the JSON/Excel report).
    if text_output:
        analyzer.save_repo_scan_results(scan_results, str(output_dir))
    
    # Ensure schema version is present for downstream validation
    if isinstance(scan_results, dict):
        scan_results.setdefault('schema_version', 'repo_scan-1.0')

    # Save JSON summary (primary location under repo tests for legacy compatibility)

    # Write advanced_analysis.json (full artifact)
    output_file = output_dir / 'advanced_analysis.json'
    with open(output_file, 'w') as f:
        json.dump(scan_results, f, indent=2)

    # Also write canonical artifact under workspace/work/advanced_analysis.json
    canonical_file = work_dir / 'advanced_analysis.json'
    with open(canonical_file, 'w') as f:
        json.dump(scan_results, f, indent=2)

    # Validate the canonical analysis artifact against schema and fail-fast on mismatch
    try:
        import sys
        from pathlib import Path as _Path

        # Ensure tool's workspace root is on sys.path so tools package can be imported
        tool_root = Path(__file__).resolve().parent.parent.parent
        if str(tool_root) not in sys.path:
            sys.path.insert(0, str(tool_root))

        from tools.schema.validate import validate_or_halt

        # Build a condensed 'repo_scan' artifact that conforms to schemas/repo_scan.schema.json
        repo_scan = {
            'schema_version': 'repo_scan-1.0',
            'repo_root': str(Path(repo_path).resolve()),
            'functions': [],
            'call_graph': [],
            'call_depth': {},
            'hardware_flags': {},
        }
        # Populate functions
        fn_index = scan_results.get('function_index', {}) if isinstance(scan_results, dict) else {}
        import hashlib
        for fid, info in fn_index.items():
            name = info.get('name') or fid
            file = info.get('file') or ''
            signature = info.get('signature') or name
            body = info.get('body') or ''
            source_hash = hashlib.sha256(body.encode('utf-8')).hexdigest() if isinstance(body, str) else ''
            repo_scan['functions'].append({
                'id': fid,
                'name': name,
                'file': file,
                'signature': signature,
                'source_hash': source_hash,
            })
        # Populate call_graph
        cg = scan_results.get('call_graph', {}) if isinstance(scan_results, dict) else {}
        if isinstance(cg, dict):
            for caller, callees in cg.items():
                if isinstance(callees, list):
                    for callee in callees:
                        repo_scan['call_graph'].append({'caller_id': caller, 'callee_id': callee})
        # call depths
        cd = scan_results.get('call_depths') or scan_results.get('call_depth') or scan_results.get('call_depths', {})
        if isinstance(cd, dict):
            for k, v in cd.items():
                repo_scan['call_depth'][k] = int(v)
        # hardware flags
        hf = scan_results.get('hardware_flags', {})
        if isinstance(hf, dict):
            repo_scan['hardware_flags'] = {k: bool(v) for k, v in hf.items()}

        # Write condensed repo_scan artifact to canonical location
        repo_scan_path = work_dir / 'repo_scan.json'
        repo_scan_path.write_text(json.dumps(repo_scan, indent=2), encoding='utf-8')

        # Use tool's schemas directory
        tool_root = Path(__file__).resolve().parent.parent.parent
        schema_path = tool_root / 'schemas' / 'repo_scan.schema.json'
        validate_or_halt(repo_scan, str(schema_path), artifact_name=repo_scan_path.name)
    except SystemExit:
        # propagate exit code from validator
        raise
    except Exception as e:
        # Any unexpected failure writing/validating should halt the pipeline
        print(f"[ERROR] Failed to validate repo_scan.json against schema: {e}")
        sys.exit(1)
    
    # Export to Excel if not disabled
    if not no_excel_output:
        excel_file = output_dir / 'analysis.xlsx'
        analyzer.export_to_excel(scan_results, str(excel_file))

    if mcdc:
        mcdc_out = output_dir / 'mcdc_gaps.json'
        payload = analyze_repo_mcdc(repo_path)
        mcdc_out.write_text(json.dumps(payload, indent=2) + "\n", encoding='utf-8')
        if not no_excel_output:
            mcdc_excel = output_dir / 'mcdc_gaps.xlsx'
            try:
                _export_mcdc_gaps_to_excel(payload, mcdc_excel)
            except Exception as e:
                if verbose:
                    print(f"[WARN] Failed to export MC/DC gaps to Excel: {e}")
        if verbose:
            print(f"MC/DC gaps saved to {mcdc_out}")
            if not no_excel_output:
                print(f"MC/DC Excel export saved to {mcdc_excel}")
    
    if verbose:
        print(f"Analysis saved to {output_file}")
        if not no_excel_output:
            print(f"Excel export saved to {excel_file}")

if __name__ == "__main__":
    main()