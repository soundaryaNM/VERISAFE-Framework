#!/usr/bin/env python3
"""
AI Test Runner - Compiles, executes, and provides coverage for AI-generated C and C++ unit tests
"""

import os
import sys
import argparse
import shutil
import subprocess
from pathlib import Path
import glob
import re
import json
import hashlib

from .safety_policy import SafetyPolicy, save_safety_summary


def _enforce_manual_review_gate(repo_root: Path) -> None:
    """MANDATORY HUMAN REVIEW GATE — DO NOT BYPASS.

    Blocks any build/run unless per-test approval flag(s) exist and match required content.
    On failure, prints the exact required message and exits non-zero.
    """

    review_dir = repo_root / "tests" / "review"
    review_required_path = review_dir / "review_required.md"

    def _approval_flag_candidates(test_path: Path) -> list[Path]:
        """Return candidate approval flag paths.

        Preferred (mirrors project structure under tests/): tests/review/<repo-path-under-tests>.flag
        Back-compat (older mirrored scheme): tests/review/<repo-relative test path>.flag
        Legacy (back-compat): tests/review/APPROVED.<filename>.flag
        """
        try:
            rel = test_path.relative_to(repo_root)
        except Exception:
            rel = Path(test_path.name)

        # Strip leading tests/ so review artifacts mirror the repo layout.
        rel_no_tests = rel
        if rel_no_tests.parts[:1] == ("tests",):
            rel_no_tests = Path(*rel_no_tests.parts[1:])

        preferred = review_dir / rel_no_tests.parent / f"{rel_no_tests.name}.flag"
        compat_mirrored = review_dir / rel.parent / f"{rel.name}.flag"
        legacy = review_dir / f"APPROVED.{test_path.name}.flag"
        return [preferred, compat_mirrored, legacy]
    def _is_iso_date(value: str) -> bool:
        import datetime
        try:
            datetime.date.fromisoformat(value)
            return True
        except Exception:
            try:
                datetime.datetime.fromisoformat(value)
                return True
            except Exception:
                return False

    def _approval_ok(content: str) -> bool:
        text = (content or "").replace("\r\n", "\n")
        lines = [ln.strip() for ln in text.split("\n") if ln.strip()]
        if len(lines) < 3:
            return False
        if lines[0].lower() != "approved = true":
            return False
        if not lines[1].lower().startswith("reviewed_by ="):
            return False
        if not lines[2].lower().startswith("date ="):
            return False

        reviewed_by = lines[1].split("=", 1)[1].strip() if "=" in lines[1] else ""
        date_val = lines[2].split("=", 1)[1].strip() if "=" in lines[2] else ""

        # Reject placeholders.
        if reviewed_by in ("", "<human_name>"):
            return False
        if date_val in ("", "<ISO date>"):
            return False
        if not _is_iso_date(date_val):
            return False
        return True

    def _parse_generated_test_files(path: Path) -> list[Path]:
        try:
            text = path.read_text(encoding="utf-8").replace("\r\n", "\n")
        except Exception:
            return []

        lines = text.split("\n")
        in_section = False
        generated: list[Path] = []

        for line in lines:
            if line.strip() == "## Generated test files":
                in_section = True
                continue
            if in_section and line.startswith("## "):
                break
            if not in_section:
                continue

            stripped = line.strip()
            if not stripped.startswith("-"):
                continue
            item = stripped.lstrip("-").strip()
            if not item or item == "(none)":
                continue

            # Normalize separators and interpret as repo-relative.
            item = item.replace("\\", "/")
            candidate = Path(item)
            if candidate.is_absolute():
                resolved = candidate
            else:
                resolved = repo_root / candidate
                if not resolved.exists():
                    alt = repo_root / "tests" / candidate
                    if alt.exists():
                        resolved = alt
            generated.append(resolved)

        return generated

    generated_test_files = _parse_generated_test_files(review_required_path)
    if not generated_test_files:
        print("❌ Manual review not approved. Build and execution halted.")
        raise SystemExit(3)

    for test_path in generated_test_files:
        content = ""
        found = False
        for approved_path in _approval_flag_candidates(test_path):
            try:
                content = approved_path.read_text(encoding="utf-8").replace("\r\n", "\n")
                found = True
                break
            except Exception:
                continue

        if not found or not _approval_ok(content):
            print("❌ Manual review not approved. Build and execution halted.")
            raise SystemExit(3)


def _cmake_test_name_for_test_file_rel(test_file_rel: str) -> str:
    """Match RailwaySignalSystem/tests/CMakeLists.txt naming scheme.

    test_file_rel is repo-relative (e.g. "tests/src/logic/test_Interlocking.cpp").
    """
    rel = (test_file_rel or "").replace("\\", "/")
    if rel.startswith("tests/"):
        rel = rel[len("tests/"):]
    p = Path(rel)
    stem = p.stem
    dir_part = str(p.parent).replace("\\", "/")
    if dir_part in ("", "."):
        return stem
    return f"{stem}__{dir_part.replace('/', '__')}"


def _suite_prefix_from_section_name(section_name: str) -> str | None:
    name = (section_name or "").strip()
    if not name:
        return None

    # Demo-safe expected: <kind>_<index> (e.g. "base_1", "mcdc_2").
    m = re.fullmatch(r"(?P<kind>[a-z_]+)_(?P<idx>\d+)", name)
    if not m:
        return None

    kind = (m.group("kind") or "").strip().upper()
    try:
        idx = int(m.group("idx"))
    except Exception:
        return None

    suffix = "" if idx == 1 else f"_{idx}"
    return f"AISEC_{kind}{suffix}_"


def _public_section_label(section: dict) -> str:
    kind = str(section.get("kind") or "base").lower()
    name = str(section.get("name") or "")
    idx = None
    m = re.fullmatch(r"[a-z_]+_(\d+)", name)
    if m:
        try:
            idx = int(m.group(1))
        except Exception:
            idx = None

    if kind == "base":
        return "BASE_TESTS"
    if kind == "mcdc":
        return f"MCDC_TESTS" + (f" (Decision_{idx})" if idx and idx > 1 else "")
    if kind == "boundary":
        return "BOUNDARY_TESTS"
    if kind == "error_path":
        return "ERROR_PATH_TESTS"
    return "GENERATED_TESTS"


def _compute_gtest_filter_for_scope(repo_root: Path, *, ctest_regex: str | None) -> str | None:
    approvals_path = repo_root / "tests" / ".approvals.json"
    if not approvals_path.exists():
        return None

    try:
        data = json.loads(approvals_path.read_text(encoding="utf-8"))
    except Exception:
        return None

    sections = (data or {}).get("sections", {})
    if not isinstance(sections, dict):
        return None

    rx = re.compile(ctest_regex) if ctest_regex else None

    patterns: list[str] = []
    for _, s in sections.items():
        if not isinstance(s, dict):
            continue
        if s.get("active") is not True:
            continue
        if s.get("approved") is not True:
            continue

        test_file_rel = s.get("test_file_rel")
        if not isinstance(test_file_rel, str) or not test_file_rel:
            continue

        if rx is not None:
            test_name = _cmake_test_name_for_test_file_rel(test_file_rel)
            if rx.search(test_name) is None:
                continue

        suite_prefix = _suite_prefix_from_section_name(str(s.get("name") or ""))
        if not suite_prefix:
            continue
        patterns.append(f"{suite_prefix}*.*")

    # Deduplicate while preserving order.
    seen: set[str] = set()
    uniq = [p for p in patterns if not (p in seen or seen.add(p))]
    return ":".join(uniq) if uniq else None


def _enforce_section_review_gate(repo_root: Path, *, ctest_regex: str | None) -> None:
    """V2 section-based approvals gate.

    If <repo>/tests/.approvals.json exists, it becomes the source of truth.
    Otherwise we fall back to the legacy per-test-file approval flags.
    """

    approvals_path = repo_root / "tests" / ".approvals.json"
    if not approvals_path.exists():
        _enforce_manual_review_gate(repo_root)
        return

    try:
        data = json.loads(approvals_path.read_text(encoding="utf-8"))
    except Exception as e:
        print("❌ Could not read approval data. Build and execution halted.")
        raise SystemExit(3)

    sections = (data or {}).get("sections", {})
    if not isinstance(sections, dict):
        print("❌ Invalid approval data. Build and execution halted.")
        raise SystemExit(3)

    active_sections: list[dict] = []
    for _, s in sections.items():
        if isinstance(s, dict) and s.get("active") is True:
            active_sections.append(s)

    # No active sections means nothing pending approval.
    if not active_sections:
        return

    header_re = re.compile(r"(?ms)^/\*\s*(?:AI-TEST-SECTION|AI-TESTGEN-SECTION)\s*\n(?P<body>.*?)\*/\s*\n")

    def _normalize_newlines(text: str) -> str:
        return (text or "").replace("\r\n", "\n").replace("\r", "\n")

    def _canonical_section_body(text: str) -> str:
        """Canonicalize section body for stable hashing.

        When sections are appended, extra blank separator lines can be inserted
        between the end of one section body and the next section header.
        Those separator newlines must not invalidate previously-approved hashes.
        """

        normalized = _normalize_newlines(text)
        stripped = normalized.rstrip()
        if not stripped:
            return ""
        return stripped + "\n"

    def _sha256_text(text: str) -> str:
        return hashlib.sha256(_normalize_newlines(text).encode("utf-8")).hexdigest()

    def _parse_section_hashes(file_text: str) -> set[str]:
        text = _normalize_newlines(file_text)
        matches = list(header_re.finditer(text))
        hashes: set[str] = set()
        for idx, m in enumerate(matches):
            body_start = m.end()
            body_end = matches[idx + 1].start() if idx + 1 < len(matches) else len(text)
            hashes.add(_sha256_text(_canonical_section_body(text[body_start:body_end])))
        return hashes

    rx = re.compile(ctest_regex) if ctest_regex else None

    for s in active_sections:
        test_file_rel = s.get("test_file_rel")
        if not isinstance(test_file_rel, str) or not test_file_rel:
            print("❌ Invalid approvals entry (missing test_file_rel). Build and execution halted.")
            raise SystemExit(3)

        # If the caller is running a subset (ctest -R), only gate the relevant test executables.
        if rx is not None:
            test_name = _cmake_test_name_for_test_file_rel(test_file_rel)
            if rx.search(test_name) is None:
                continue

        # Strict gate: if any ACTIVE section is unapproved (in-scope), do not start a build.
        if s.get("approved") is not True:
            print("❌ Pending section approvals exist (active). Build and execution halted.")
            raise SystemExit(3)

        section_sha = s.get("section_sha256")
        if not test_file_rel or not section_sha:
            print("❌ Invalid approval entry (missing required data). Build and execution halted.")
            raise SystemExit(3)

        test_file = (repo_root / test_file_rel).resolve()
        try:
            text = test_file.read_text(encoding="utf-8")
        except Exception:
            print(f"❌ Could not read approved test file: {test_file_rel}. Build and execution halted.")
            raise SystemExit(3)

        hashes = _parse_section_hashes(text)
        if section_sha not in hashes:
            label = _public_section_label(s)
            print(
                "❌ Approved content no longer matches the current test file. "
                "Please re-approve the current content before building."
            )
            if label:
                print(f"   Section: {label}")
            print(f"   File: {test_file_rel}")
            raise SystemExit(3)


def _select_test_files_from_v2_approvals(repo_root: Path, *, ctest_regex: str | None) -> list[Path]:
    """Best-effort selection of test files based on tests/.approvals.json.

    When ctest_regex is provided, only includes test files whose CTest name matches.
    Only ACTIVE+APPROVED sections contribute.
    """

    approvals_path = repo_root / "tests" / ".approvals.json"
    if not approvals_path.exists():
        return []

    try:
        data = json.loads(approvals_path.read_text(encoding="utf-8"))
    except Exception:
        return []

    sections = (data or {}).get("sections", {})
    if not isinstance(sections, dict):
        return []

    rx = re.compile(ctest_regex) if ctest_regex else None

    files: list[Path] = []
    seen: set[Path] = set()
    for s in sections.values():
        if not isinstance(s, dict) or s.get("active") is not True:
            continue
        if s.get("approved") is not True:
            continue

        test_file_rel = s.get("test_file_rel")
        if not isinstance(test_file_rel, str) or not test_file_rel:
            continue

        if rx is not None:
            test_name = _cmake_test_name_for_test_file_rel(test_file_rel)
            if rx.search(test_name) is None:
                continue

        p = (repo_root / test_file_rel).resolve()
        if p.exists() and p not in seen:
            seen.add(p)
            files.append(p)

    return files




class AITestRunner:
    """AI Test Runner - Builds, executes, and covers AI-generated C and C++ tests"""

    def __init__(self, repo_path: str, output_dir: str = "build", language: str = "auto"):
        self.repo_path = Path(repo_path).resolve()
        out = Path(output_dir)
        if out.is_absolute():
            self.output_dir = out
        else:
            # Enforce build output under <repo>/tests/ to avoid separate top-level build folders.
            # Examples:
            #   output_dir=build         -> <repo>/tests/build
            #   output_dir=tests/build   -> <repo>/tests/build
            #   output_dir=ai_test_build -> <repo>/tests/ai_test_build
            if out.parts[:1] == ("tests",):
                self.output_dir = self.repo_path / out
            else:
                self.output_dir = self.repo_path / "tests" / out
        self.tests_dir = self.repo_path / "tests"
        self.verification_dir = self.tests_dir / "compilation_report"
        self.test_reports_root = self.tests_dir / "test_reports"
        self.source_dir = self.repo_path / "src"
        self.language = language  # "c", "cpp", or "auto"
        self.ctest_regex: str | None = None
        self.report_group: str = "all"
        import xml.etree.ElementTree as ET

        # Create output directory
        self.output_dir.mkdir(parents=True, exist_ok=True)
        # Create test reports directory
        self.test_reports_root.mkdir(parents=True, exist_ok=True)

    def _sanitize_group_name(self, name: str) -> str:
        safe = re.sub(r"[^A-Za-z0-9_.-]+", "_", (name or "").strip())
        return safe or "all"

    def _derive_report_group(self) -> str:
        if self.ctest_regex:
            # When demo selects a single file it passes the stem; when multiple it passes `a|b|c`.
            parts = [p.strip() for p in self.ctest_regex.split("|") if p.strip()]
            if len(parts) == 1:
                return self._sanitize_group_name(parts[0])
            return "selected"
        return "all"

    def _report_dir(self) -> Path:
        # Keep `report_group` as metadata for summaries, but do not encode it into
        # the folder structure. Users expect reports to mirror source/test paths:
        #   tests/test_reports/src/<...>/<file>
        # not:
        #   tests/test_reports/<regex-or-group>/src/<...>/<file>
        self.report_group = self._derive_report_group()
        self.test_reports_root.mkdir(parents=True, exist_ok=True)

        # One-time best-effort migration: older versions wrote reports under
        # tests/test_reports/<group>/src/.... Move those into tests/test_reports/src/...
        # so the on-disk layout is consistent and doesn't confuse users.
        try:
            legacy_dirs = [p for p in self.test_reports_root.iterdir() if p.is_dir() and p.name != "src"]
            for legacy in legacy_dirs:
                legacy_src = legacy / "src"
                if not legacy_src.exists() or not legacy_src.is_dir():
                    continue

                target_src = self.test_reports_root / "src"
                target_src.mkdir(parents=True, exist_ok=True)

                for item in legacy_src.rglob("*"):
                    if item.is_dir():
                        continue
                    rel = item.relative_to(legacy_src)
                    dest = target_src / rel
                    dest.parent.mkdir(parents=True, exist_ok=True)
                    if dest.exists():
                        # Preserve existing; keep legacy file with a suffix.
                        dest = dest.with_name(dest.stem + "_legacy" + dest.suffix)
                    try:
                        shutil.move(str(item), str(dest))
                    except Exception:
                        pass

                # Remove legacy folder if it's empty after move.
                try:
                    def _on_rm_error(func, path, exc_info):
                        try:
                            os.chmod(path, 0o700)
                            func(path)
                        except Exception:
                            pass

                    shutil.rmtree(legacy, onerror=_on_rm_error)
                except Exception:
                    pass
        except Exception:
            pass
        return self.test_reports_root

    def _clean_report_root(self, report_root: Path) -> None:
        """Normalize report layout so tests/test_reports stays readable.

        Policy:
        - Only keep a single `src/` tree plus a few top-level index files.
        - Remove legacy/nested folders and previous-run artifacts.
        """
        report_root.mkdir(parents=True, exist_ok=True)

        # Defensive: if an older layout accidentally created tests/test_reports/test_reports/...
        nested = report_root / "test_reports"
        if nested.exists() and nested.is_dir():
            try:
                shutil.rmtree(nested)
            except Exception:
                pass

        # Remove prior run index files to keep the root tidy.
        # Remove prior run index files to keep the root tidy.
        for legacy_name in ("SUMMARY.txt", "RESULTS.csv", "RESULTS.xlsx"):
            p = report_root / legacy_name
            if p.exists() and p.is_file():
                try:
                    p.unlink()
                except Exception:
                    pass

        # Remove any leftover index folder from older runs.
        index_dir = report_root / "index"
        if index_dir.exists() and index_dir.is_dir():
            try:
                shutil.rmtree(index_dir)
            except Exception:
                pass

        # Wipe the mirrored tree every run to avoid accumulating legacy files.
        src_dir = report_root / "src"
        if src_dir.exists() and src_dir.is_dir():
            try:
                shutil.rmtree(src_dir)
            except Exception:
                pass
        try:
            src_dir.mkdir(parents=True, exist_ok=True)
        except Exception:
            pass

    def detect_language(self, test_files):
        """Detect the programming language from test files"""
        if self.language != "auto":
            return self.language

        # Check file extensions
        cpp_extensions = ['.cpp', '.cc', '.cxx', '.c++']
        c_extensions = ['.c']

        has_cpp = any(any(test_file.name.endswith(ext) for ext in cpp_extensions) for test_file in test_files)
        has_c = any(any(test_file.name.endswith(ext) for ext in c_extensions) for test_file in test_files)

        if has_cpp:
            return "cpp"
        elif has_c:
            return "c"
        else:
            return "cpp"  # Default to C++

    def find_compilable_tests(self):
        """Find test files that have compiles_yes in verification reports"""
        print("Starting find_compilable_tests")
        compilable_tests = []

        if not self.verification_dir.exists():
            print(f"⚠️  Verification report directory not found: {self.verification_dir}")
            print("⚠️  Using approved test list to select tests.")
            return _select_test_files_from_v2_approvals(self.repo_path, ctest_regex=getattr(self, 'ctest_regex', None))

        # Find all compiles_yes files (recursive; reports mirror test folder structure)
        for report_file in self.verification_dir.rglob("*compiles_yes.txt"):
            rel_report = report_file.relative_to(self.verification_dir)
            base_name = report_file.name.replace("_compiles_yes.txt", "")

            # Try both .c and .cpp extensions under the same mirrored subfolder.
            for ext in ['.cpp', '.cc', '.cxx', '.c++', '.c']:
                test_file = self.tests_dir / rel_report.parent / f"{base_name}{ext}"
                if test_file.exists():
                    compilable_tests.append(test_file)
                    print(f"✅ Found compilable test: {test_file}")
                    break

        if not compilable_tests:
            print("⚠️  No verified compilable tests found.")
            print("⚠️  Using approved test list to select tests.")
            return _select_test_files_from_v2_approvals(self.repo_path, ctest_regex=getattr(self, 'ctest_regex', None))

        return compilable_tests

    def copy_unity_framework(self):
        """Copy or download Unity framework"""
        unity_dest = self.output_dir / "unity"

        # First try to copy from reference location
        unity_source = self.repo_path.parent / "ai-test-gemini-CLI" / "unity"
        if unity_source.exists() and any(unity_source.rglob("*.c")):
            if unity_dest.exists():
                try:
                    shutil.rmtree(unity_dest)
                except (OSError, PermissionError):
                    print(f"⚠️  Could not remove existing unity directory: {unity_dest}")
            shutil.copytree(unity_source, unity_dest)
            print("✅ Copied Unity framework from reference")
            return True

        # If not available, download Unity
        print("📥 Downloading Unity framework...")
        import urllib.request
        import zipfile
        import tempfile

        try:
            # Download Unity from GitHub
            unity_url = "https://github.com/ThrowTheSwitch/Unity/archive/refs/heads/master.zip"
            with tempfile.NamedTemporaryFile(suffix='.zip', delete=False) as temp_zip:
                temp_zip_path = temp_zip.name

            # Download to the temp file
            urllib.request.urlretrieve(unity_url, temp_zip_path)

            # Extract Unity
            with zipfile.ZipFile(temp_zip_path, 'r') as zip_ref:
                # Extract only the src directory
                for member in zip_ref.namelist():
                    if member.startswith('Unity-master/src/'):
                        # Remove the Unity-master/src/ prefix
                        target_path = member.replace('Unity-master/src/', 'src/')
                        if target_path.endswith('/'):
                            (unity_dest / target_path).mkdir(parents=True, exist_ok=True)
                        else:
                            zip_ref.extract(member, unity_dest.parent / "temp_unity")
                            source_file = unity_dest.parent / "temp_unity" / member
                            target_file = unity_dest / target_path
                            target_file.parent.mkdir(parents=True, exist_ok=True)
                            shutil.move(source_file, target_file)

            # Clean up
            import os
            os.unlink(temp_zip_path)
            temp_dir = unity_dest.parent / "temp_unity"
            if temp_dir.exists():
                shutil.rmtree(temp_dir)

            print("✅ Downloaded Unity framework")
            return True

        except Exception as e:
            print(f"❌ Failed to download Unity: {e}")
            print("⚠️  Unity framework not available, tests may not compile")
            return False

    def setup_cpp_framework(self):
        """Setup C++ test framework (Google Test) and Arduino stubs"""
        print("📦 Setting up C++ test framework...")

        # Copy Google Test header
        gtest_dest = self.output_dir / "gtest" / "gtest.h"
        gtest_dest.parent.mkdir(parents=True, exist_ok=True)

        # Try to copy from reference location first
        gtest_source = self.repo_path.parent / "Door-Monitoring" / "tests_and_build_single_file" / "gtest" / "gtest.h"
        if gtest_source.exists():
            shutil.copy2(gtest_source, gtest_dest)
            print("✅ Copied Google Test header from reference")
        else:
            # Create minimal Google Test framework
            gtest_content = '''#pragma once
// Minimal Google Test-like framework for testing
#include <iostream>
#include <vector>
#include <functional>
#include <cassert>

class TestRegistry {
private:
    struct TestInfo {
        std::string name;
        std::function<void()> func;
    };
    std::vector<TestInfo> tests_;
    static TestRegistry* instance_;

    TestRegistry() {}

public:
    static TestRegistry& instance() {
        if (!instance_) instance_ = new TestRegistry();
        return *instance_;
    }

    void register_test(const std::string& name, std::function<void()> func) {
        tests_.push_back({name, func});
    }

    int run_all_tests() {
        int failures = 0;
        for (const auto& test : tests_) {
            try {
                test.func();
                std::cout << "[ PASS ] " << test.name << std::endl;
            } catch (const std::exception& e) {
                std::cout << "[ FAIL ] " << test.name << ": " << e.what() << std::endl;
                failures++;
            } catch (...) {
                std::cout << "[ FAIL ] " << test.name << ": Unknown exception" << std::endl;
                failures++;
            }
        }
        return failures;
    }
};

TestRegistry* TestRegistry::instance_ = nullptr;

#define TEST(suite, name) \\
    void Test_##suite##_##name(); \\
    struct Registrar_##suite##_##name { \\
        Registrar_##suite##_##name() { \\
            TestRegistry::instance().register_test(#suite "." #name, Test_##suite##_##name); \\
        } \\
    } registrar_##suite##_##name; \\
    void Test_##suite##_##name()

#define ASSERT_EQ(a, b) assert((a) == (b))
#define ASSERT_NE(a, b) assert((a) != (b))
#define ASSERT_TRUE(a) assert((a))
#define ASSERT_FALSE(a) assert(!(a))

int RUN_ALL_TESTS() {
    return TestRegistry::instance().run_all_tests();
}
'''
            with open(gtest_dest, 'w') as f:
                f.write(gtest_content)
            print("✅ Created minimal Google Test framework")

        # Copy Arduino stubs
        arduino_dest = self.output_dir / "arduino_stubs"
        arduino_dest.mkdir(parents=True, exist_ok=True)

        # Try to copy from reference location
        arduino_source = self.repo_path.parent / "Door-Monitoring" / "tests_and_build_single_file"
        stubs_files = ["Arduino_stubs.h", "Arduino_stubs.cpp"]
        copied = False

        for stub_file in stubs_files:
            source_file = arduino_source / stub_file
            if source_file.exists():
                shutil.copy2(source_file, arduino_dest / stub_file)
                copied = True

        if copied:
            print("✅ Copied Arduino stubs from reference")
        else:
            # Create Arduino stubs with expected globals for testing
            arduino_h_content = '''#pragma once

#include <string>
#include <vector>
#include <iostream>
#include <chrono>
#include <thread>

void digitalWrite(int pin, int value);
int digitalRead(int pin);
void pinMode(int pin, int mode);
void delay(int ms);
unsigned long millis();
void reset_arduino_stubs();

class String {
private:
    std::string data;

public:
    String();
    String(const char* str);
    String(int val);
    String& operator+=(const char* str);
    String operator+(const char* str) const;
    String operator+(const String& other) const;
    const char* c_str() const;
    
    friend String operator+(const char* lhs, const String& rhs);
};

struct DigitalWriteCall {
    int pin;
    int value;
};

struct DelayCall {
    int ms;
};

class SerialClass {
public:
    void begin(int baud);
    void print(const char* str);
    void println(const char* str);
    void print(int val);
    void println(int val);
    void print(const String& str);
    void println(const String& str);
    
    int begin_call_count = 0;
    int last_baud_rate = 0;
    int println_call_count = 0;
    int print_call_count = 0;
    
    std::string outputBuffer;
};

extern SerialClass Serial;
extern std::vector<DigitalWriteCall> digitalWrite_calls;
extern std::vector<DelayCall> delay_calls;

#define HIGH 1
#define LOW 0
#define INPUT 0
#define OUTPUT 1
#define LED 13
'''

            arduino_cpp_content = '''#include "Arduino_stubs.h"
#include <iostream>
#include <map>
#include <chrono>

static std::map<int, int> pin_states;
static auto start_time = std::chrono::steady_clock::now();

std::vector<DigitalWriteCall> digitalWrite_calls;
std::vector<DelayCall> delay_calls;

void reset_arduino_stubs() {
    Serial.begin_call_count = 0;
    Serial.last_baud_rate = 0;
    Serial.println_call_count = 0;
    Serial.print_call_count = 0;
    digitalWrite_calls.clear();
    delay_calls.clear();
    Serial.outputBuffer.clear();
    pin_states.clear();
}

void digitalWrite(int pin, int value) {
    pin_states[pin] = value;
    digitalWrite_calls.push_back({pin, value});
}

int digitalRead(int pin) {
    return pin_states[pin];
}

void pinMode(int pin, int mode) {
    // Not tracked for testing
}

void delay(int ms) {
    delay_calls.push_back({ms});
    std::this_thread::sleep_for(std::chrono::milliseconds(ms));
}

unsigned long millis() {
    auto now = std::chrono::steady_clock::now();
    auto duration = now - start_time;
    return std::chrono::duration_cast<std::chrono::milliseconds>(duration).count();
}

SerialClass Serial;

void SerialClass::begin(int baud) {
    begin_call_count++;
    last_baud_rate = baud;
}

void SerialClass::print(const char* str) {
    print_call_count++;
    outputBuffer += str;
}

void SerialClass::println(const char* str) {
    println_call_count++;
    outputBuffer += str;
    outputBuffer += "\\n";
}

void SerialClass::print(int val) {
    print_call_count++;
    outputBuffer += std::to_string(val);
}

void SerialClass::println(int val) {
    println_call_count++;
    outputBuffer += std::to_string(val);
    outputBuffer += "\\n";
}

void SerialClass::print(const String& str) {
    print_call_count++;
    outputBuffer += str.c_str();
}

void SerialClass::println(const String& str) {
    println_call_count++;
    outputBuffer += str.c_str();
    outputBuffer += "\\n";
}

String::String() {}

String::String(const char* str) : data(str) {}

String::String(int val) : data(std::to_string(val)) {}

String& String::operator+=(const char* str) {
    data += str;
    return *this;
}

String String::operator+(const char* str) const {
    String result = *this;
    result.data += str;
    return result;
}

String String::operator+(const String& other) const {
    String result = *this;
    result.data += other.data;
    return result;
}

String operator+(const char* lhs, const String& rhs) {
    String result(lhs);
    result.data += rhs.data;
    return result;
}

const char* String::c_str() const {
    return data.c_str();
}
'''
            with open(arduino_dest / "Arduino_stubs.h", 'w') as f:
                f.write(arduino_h_content)
            with open(arduino_dest / "Arduino_stubs.cpp", 'w') as f:
                f.write(arduino_cpp_content)
            print("✅ Created minimal Arduino stubs")

        return True

    def copy_source_files(self):
        """Copy source files to build directory and generate headers"""
        src_build_dir = self.output_dir / "src"
        src_build_dir.mkdir(exist_ok=True)

        if self.source_dir.exists():
            # Copy C and C++ source files
            source_files = list(self.source_dir.glob("*.c")) + list(self.source_dir.glob("*.cpp"))
            
            for src_file in source_files:
                with open(src_file, 'r') as f:
                    content = f.read()
                
                # Rename main() to app_main() to allow testing it without conflicts
                import re
                if 'int main' in content:
                    content = re.sub(r'\bint\s+main\s*\(', 'int app_main(', content)
                    print(f"🔄 Renamed main() to app_main() in {src_file.name}")
                
                # Write to build directory
                dest_file = src_build_dir / src_file.name
                with open(dest_file, 'w') as f:
                    f.write(content)
                print(f"📋 Copied source: {src_file.name}")
                
                # Generate a header file (only for C files usually, but maybe useful for CPP too if missing)
                if src_file.suffix == '.c':
                    header_file = src_build_dir / (src_file.stem + ".h")
                    self._generate_header_from_source(src_file, header_file)

            for header_file in self.source_dir.glob("*.h"):
                shutil.copy2(header_file, src_build_dir)
                print(f"📋 Copied header: {header_file.name}")
        else:
            print(f"⚠️  Source directory not found: {self.source_dir}")

    def _generate_header_from_source(self, src_file, dest_header):
        """Generate a header file from source with function declarations"""
        try:
            with open(src_file, 'r') as f:
                content = f.read()
            
            # Extract function definitions (anything that looks like a function)
            import re
            # Match patterns like: return_type function_name(parameters) {
            pattern = r'(\w+\s+(\w+)\s*\([^)]*\))\s*\{'
            matches = re.findall(pattern, content)
            
            if matches:
                with open(dest_header, 'w') as f:
                    f.write(f"/* Auto-generated header for {src_file.name} */\n")
                    f.write("#pragma once\n\n")
                    f.write("#include <stdint.h>\n")
                    f.write("#include <stdbool.h>\n")
                    f.write("#include <stdlib.h>\n\n")
                    
                    for match in matches:
                        func_name = match[1]
                        func_decl = match[0]
                        # Skip main function
                        if func_name != 'main':
                            f.write(f"{func_decl};\n")
                print(f"📝 Generated header: {dest_header.name}")
        except Exception as e:
            print(f"⚠️  Could not generate header: {e}")

    def copy_test_files(self, test_files):
        """Copy test files to build directory"""
        tests_build_dir = self.output_dir / "tests"
        tests_build_dir.mkdir(exist_ok=True)
        import re

        for test_file in test_files:
            with open(test_file, 'r') as f:
                content = f.read()
            
            # For C tests, inject #include for the source C file to get implementations
            if test_file.name.endswith('.c') and '#include "unity.h"' in content:
                # Determine source filename
                test_name = test_file.stem
                if test_name.startswith('test_'):
                    source_name = test_name[5:]
                else:
                    source_name = test_name
                
                # Include the actual source C file (not header)
                # Since test is in tests/ and source is in src/, use ../src/filename.c
                source_file = f"../src/{source_name}.c"
                if f'#include "{source_file}"' not in content and f'#include "../src/' not in content:
                    # Find where to insert (after unity.h include)
                    unity_pos = content.find('#include "unity.h"')
                    if unity_pos != -1:
                        # Find the end of the line
                        eol = content.find('\n', unity_pos) + 1
                        # Insert the include
                        content = content[:eol] + f'#include "{source_file}"\n' + content[eol:]
                        print(f"📝 Added #include for {source_file} to test file")
                
                # Replace calls to main() with app_main() to avoid recursion
                # But don't replace the test runner's main definition (int main(void))
                if 'app_main(' not in content:
                    # Regex: match 'main(' not preceded by 'int ' or 'void '
                    new_content = re.sub(r'(?<!\bint\s)(?<!\bvoid\s)\bmain\s*\(', 'app_main(', content)
                    if new_content != content:
                        content = new_content
                        print(f"🔄 Replaced main() calls with app_main() in {test_file.name}")
            
            # Write the modified test file
            dest_file = tests_build_dir / test_file.name
            with open(dest_file, 'w') as f:
                f.write(content)
            print(f"📋 Copied test: {test_file.name}")

    def create_cmake_lists(self, test_files, language):
        """Create CMakeLists.txt based on language"""
        if language == "cpp":
            return self.create_cpp_cmake_lists(test_files)
        else:
            return self.create_c_cmake_lists(test_files)

    def create_c_cmake_lists(self, test_files):
        """Create CMakeLists.txt for C tests with Unity"""
        cmake_content = "cmake_minimum_required(VERSION 3.10)\n"
        cmake_content += "project(Tests C)\n\n"
        cmake_content += "set(CMAKE_C_STANDARD 99)\n"
        cmake_content += "add_definitions(-DUNIT_TEST)\n\n"
        cmake_content += "set(CMAKE_C_FLAGS \"${CMAKE_C_FLAGS} --coverage\")\n"
        cmake_content += "set(CMAKE_EXE_LINKER_FLAGS \"${CMAKE_EXE_LINKER_FLAGS} --coverage\")\n\n"
        cmake_content += "include_directories(unity/src)\n"
        cmake_content += "include_directories(src)\n\n"
        cmake_content += "add_library(unity unity/src/unity.c)\n\n"

        for test_file in test_files:
            test_name = os.path.splitext(os.path.basename(test_file))[0]
            executable_name = test_name

            # For C tests with Unity, only compile the test file
            # The test file should include the source file to get function definitions
            test_file_basename = os.path.basename(test_file).replace('\\', '/')
            cmake_content += f"add_executable({executable_name} tests/{test_file_basename})\n"
            cmake_content += f"target_link_libraries({executable_name} unity)\n\n"

        with open(os.path.join(self.output_dir, 'CMakeLists.txt'), 'w') as f:
            f.write(cmake_content)
        print(f"✅ Created CMakeLists.txt for C tests with {len(test_files)} targets")
        return True

    def create_cpp_cmake_lists(self, test_files):
        """Create CMakeLists.txt for C++ tests with Google Test"""
        cmake_content = "cmake_minimum_required(VERSION 3.14)\n"
        cmake_content += "project(cpp_tests CXX)\n\n"
        cmake_content += "set(CMAKE_CXX_STANDARD 17)\n"
        cmake_content += "set(CMAKE_CXX_STANDARD_REQUIRED ON)\n\n"
        cmake_content += "enable_testing()\n\n"

        # Add source files under test
        source_files = []
        if self.source_dir.exists():
            for ext in ['.cpp', '.cc', '.cxx', '.c++']:
                source_files.extend(self.source_dir.glob(f"*{ext}"))

        if source_files:
            cmake_content += "# Source code under test\n"
            cmake_content += "add_library(test_lib OBJECT\n"
            for src_file in source_files:
                cmake_content += f"  src/{src_file.name}\n"
            cmake_content += ")\n"
            cmake_content += "target_include_directories(test_lib PUBLIC ${CMAKE_CURRENT_SOURCE_DIR}/src)\n"
            cmake_content += "target_include_directories(test_lib PUBLIC arduino_stubs)\n"
            cmake_content += "target_include_directories(test_lib PUBLIC gtest)\n\n"

        # Add test executables
        for test_file in test_files:
            test_name = test_file.stem
            cmake_content += f"# Test executable for {test_name}\n"
            cmake_content += f"add_executable({test_name}\n"
            cmake_content += f"  tests/{test_file.name}\n"
            cmake_content += f"  arduino_stubs/Arduino_stubs.cpp\n"
            if source_files:
                cmake_content += f"  $<TARGET_OBJECTS:test_lib>\n"
            cmake_content += ")\n"
            cmake_content += f"target_include_directories({test_name} PRIVATE ${{CMAKE_CURRENT_SOURCE_DIR}})\n"
            cmake_content += f"target_include_directories({test_name} PRIVATE ${{CMAKE_CURRENT_SOURCE_DIR}}/src)\n"
            cmake_content += f"target_include_directories({test_name} PRIVATE arduino_stubs)\n"
            cmake_content += f"target_include_directories({test_name} PRIVATE gtest)\n\n"
            cmake_content += f"add_test(\n"
            cmake_content += f"  NAME {test_name}\n"
            cmake_content += f"  COMMAND {test_name}\n"
            cmake_content += ")\n\n"

        # Write CMakeLists.txt
        cmake_file = self.output_dir / "CMakeLists.txt"
        with open(cmake_file, 'w') as f:
            f.write(cmake_content)

        print("✅ Created CMakeLists.txt for C++ tests")
        return True

    def build_tests(self):
        """Build the tests using CMake"""
        print("🔨 Building tests...")

        try:
            # Configure with CMake
            result = subprocess.run(
                ["cmake", "."],
                cwd=self.output_dir,
                capture_output=True,
                text=True,
                check=True
            )
            print("✅ CMake configuration successful")

            # Build with cmake --build
            result = subprocess.run(
                ["cmake", "--build", "."],
                cwd=self.output_dir,
                capture_output=True,
                text=True,
                check=True
            )
            print("✅ Build successful")

        except subprocess.CalledProcessError as e:
            print(f"❌ Build failed: {e}")
            print(f"STDOUT: {e.stdout}")
            print(f"STDERR: {e.stderr}")
            return False
        except FileNotFoundError:
            print("❌ CMake not found. Please install CMake.")
            return False

        return True

    def run_tests(self):
        """Run the compiled tests"""
        print("🧪 Running tests...")

        test_results = []
        test_executables = [exe for exe in self.output_dir.glob("*test*") 
                           if exe.is_file() and exe.suffix in ['.exe', ''] and 'CTest' not in exe.name]

        if not test_executables:
            print("❌ No test executables found")
            return test_results

        for exe in test_executables:
            if exe.is_file() and os.access(exe, os.X_OK):
                print(f"   Running {exe.name}...")
                try:
                    result = subprocess.run(
                        [str(exe)],
                        cwd=self.output_dir,
                        capture_output=True,
                        text=True,
                        timeout=30
                    )

                    # Parse test output
                    individual_tests = 0
                    individual_passed = 0
                    individual_failed = 0

                    for line in result.stdout.split('\n'):
                        line = line.strip()
                        if ':PASS' in line:
                            individual_tests += 1
                            individual_passed += 1
                        elif ':FAIL' in line:
                            individual_tests += 1
                            individual_failed += 1
                        elif line.endswith('Tests') and 'Failures' in line:
                            parts = line.split()
                            if len(parts) >= 3:
                                try:
                                    individual_tests = int(parts[0])
                                    individual_failed = int(parts[2])
                                    individual_passed = individual_tests - individual_failed
                                except ValueError:
                                    pass

                    success = result.returncode == 0
                    test_results.append({
                        'name': exe.name,
                        'success': success,
                        'output': result.stdout,
                        'errors': result.stderr,
                        'returncode': result.returncode,
                        'individual_tests': individual_tests,
                        'individual_passed': individual_passed,
                        'individual_failed': individual_failed
                    })

                    status = "✅" if success else "❌"
                    if individual_tests > 0:
                        print(f"   {status} {exe.name} ({individual_passed}/{individual_tests} tests passed)")
                    else:
                        print(f"   {status} {exe.name} (exit code: {result.returncode})")

                except subprocess.TimeoutExpired:
                    test_results.append({
                        'name': exe.name,
                        'success': False,
                        'output': '',
                        'errors': 'Test timed out',
                        'returncode': -1,
                        'individual_tests': 0,
                        'individual_passed': 0,
                        'individual_failed': 0
                    })
                    print(f"   ⏰ {exe.name} timed out")

                except Exception as e:
                    test_results.append({
                        'name': exe.name,
                        'success': False,
                        'output': '',
                        'errors': str(e),
                        'returncode': -1,
                        'individual_tests': 0,
                        'individual_passed': 0,
                        'individual_failed': 0
                    })
                    print(f"   ❌ {exe.name} failed: {e}")

        return test_results

    def generate_test_reports(self, test_results):
        """Generate individual test reports"""
        print(f"📝 Generating test reports in {self.test_reports_dir}...")

        for result in test_results:
            report_file = self.test_reports_dir / f"{result['name']}_report.txt"

            with open(report_file, 'w', encoding='utf-8') as f:
                f.write("=" * 60 + "\n")
                f.write(f"TEST REPORT: {result['name']}\n")
                f.write("=" * 60 + "\n\n")

                f.write("EXECUTION SUMMARY\n")
                f.write("-" * 20 + "\n")
                f.write(f"Test Executable: {result['name']}\n")
                f.write(f"Exit Code: {result['returncode']}\n")
                f.write(f"Overall Status: {'PASSED' if result['success'] else 'FAILED'}\n")
                f.write(f"Individual Tests Run: {result['individual_tests']}\n")
                f.write(f"Individual Tests Passed: {result['individual_passed']}\n")
                f.write(f"Individual Tests Failed: {result['individual_failed']}\n\n")

                if result['errors']:
                    f.write("ERRORS\n")
                    f.write("-" * 10 + "\n")
                    f.write(f"{result['errors']}\n\n")

                f.write("DETAILED OUTPUT\n")
                f.write("-" * 20 + "\n")
                if result['output']:
                    f.write(result['output'])
                else:
                    f.write("(No output captured)\n")

                f.write("\n" + "=" * 60 + "\n")

            print(f"   📄 Generated report: {report_file.name}")
            # Also show the relative path for easier navigation.
            try:
                rel_print = report_file.relative_to(self.repo_path).as_posix()
                print(f"      ↳ {rel_print}")
            except Exception:
                pass

    def generate_coverage(self):
        """Generate coverage reports (placeholder)"""
        print("📊 Coverage reporting not yet implemented")
        return True

    def create_cmake_lists(self, test_files, language):
        """Create CMakeLists.txt in repo root for CMake build"""
        cmake_path = self.repo_path / "CMakeLists.txt"
        with open(cmake_path, 'w') as f:
            f.write("cmake_minimum_required(VERSION 3.14)\n")
            f.write("project(TestProject CXX)\n\n")
            
            f.write("set(CMAKE_CXX_STANDARD 17)\n")
            f.write("set(CMAKE_CXX_STANDARD_REQUIRED ON)\n\n")
            
            f.write("# Enable testing\n")
            f.write("enable_testing()\n\n")
            
            if language == "cpp":
                f.write("# Google Test setup\n")
                # Disable SSL verification for download to avoid certificate issues
                f.write("set(CMAKE_TLS_VERIFY 0)\n")
                f.write("include(FetchContent)\n")
                f.write("FetchContent_Declare(\n")
                f.write("  googletest\n")
                f.write("  URL https://github.com/google/googletest/archive/refs/tags/v1.14.0.zip\n")
                f.write(")\n")
                f.write("# For Windows: Prevent overriding the parent project's compiler/linker settings\n")
                f.write("set(gtest_force_shared_crt ON CACHE BOOL \"\" FORCE)\n")
                f.write("FetchContent_MakeAvailable(googletest)\n\n")
                
                f.write("include_directories(${gtest_SOURCE_DIR}/include)\n\n")

            for test_file in test_files:
                # Assume test files are in tests/ directory
                exe_name = test_file.stem
                
                # Determine source file name (assuming test_X.cpp tests X.cpp)
                source_name = test_file.stem
                if source_name.startswith("test_"):
                    source_name = source_name[5:]
                
                # Look for source file in src/ or root
                source_file_path = self.repo_path / "src" / f"{source_name}.cpp"
                if not source_file_path.exists():
                     # Try root
                     source_file_path = self.repo_path / f"{source_name}.cpp"
                
                # If source file exists, create a library for it
                if source_file_path.exists():
                    lib_name = f"{source_name}_lib"
                    # Use relative path for CMake
                    rel_source_path = source_file_path.relative_to(self.repo_path).as_posix()
                    
                    f.write(f"add_library({lib_name} OBJECT {rel_source_path})\n")
                    f.write(f"target_include_directories({lib_name} PUBLIC ${{CMAKE_CURRENT_SOURCE_DIR}})\n")
                    f.write(f"target_include_directories({lib_name} PUBLIC ${{CMAKE_BINARY_DIR}}/arduino_stubs)\n\n")
                    
                    f.write(f"add_executable({exe_name} tests/{test_file.name} ${{CMAKE_BINARY_DIR}}/arduino_stubs/Arduino_stubs.cpp $<TARGET_OBJECTS:{lib_name}>)\n")
                else:
                    # Fallback: just compile test file (might fail linking)
                    f.write(f"add_executable({exe_name} tests/{test_file.name} ${{CMAKE_BINARY_DIR}}/arduino_stubs/Arduino_stubs.cpp)\n")

                if language == "cpp":
                    f.write(f"target_link_libraries({exe_name} GTest::gtest_main)\n")
                
                f.write(f"target_include_directories({exe_name} PRIVATE ${{CMAKE_CURRENT_SOURCE_DIR}} ${{CMAKE_CURRENT_SOURCE_DIR}}/src ${{CMAKE_BINARY_DIR}}/arduino_stubs)\n")
                
                f.write(f"add_test(NAME {exe_name} COMMAND {exe_name})\n")
                
        print(f"📝 Created CMakeLists.txt at {cmake_path}")
        return True

    def run(self):
        """Run the complete test execution pipeline"""
        print("🚀 Starting AI Test Runner...")

        # Find compilable tests
        test_files = self.find_compilable_tests()
        if not test_files:
            print("❌ No compilable tests found")
            return False

        # Detect language
        language = self.detect_language(test_files)
        print(f"🔍 Detected language: {language.upper()}")

        # Setup test framework based on language
        if language == "cpp":
            if not self.setup_cpp_framework():
                print("❌ Failed to setup C++ test framework")
                return False
        else:  # C
            if not self.copy_unity_framework():
                print("❌ Failed to setup Unity framework")
                return False

        # Copy source and test files
        self.copy_source_files()
        self.copy_test_files(test_files)

        # Create CMakeLists.txt
        # self.create_cmake_lists(test_files, language)  # Project already has CMakeLists.txt

        # Configure CMake
        print("🔧 Configuring CMake...")
        try:
            cmake_args = [
                "cmake",
                "-S",
                str(self.repo_path),
                "-B",
                str(self.output_dir),
                "-DRAILWAY_FETCH_GTEST=ON",
            ]

            # Compatibility: this demo repo uses a namespaced option.
            try:
                top_cmake = (self.repo_path / "CMakeLists.txt").read_text(encoding="utf-8", errors="ignore")
                # Generic coverage flag for any repo that supports it.
                # Use a word-boundary-ish regex so we don't match namespaced options.
                if re.search(r"(?<![A-Z0-9_])ENABLE_COVERAGE(?![A-Z0-9_])", top_cmake):
                    cmake_args.append("-DENABLE_COVERAGE=ON")
                if "RAILWAY_ENABLE_COVERAGE" in top_cmake:
                    cmake_args.append("-DRAILWAY_ENABLE_COVERAGE=ON")
            except Exception:
                pass

            subprocess.run(cmake_args, check=True)
        except subprocess.CalledProcessError as e:
            print(f"❌ CMake configuration failed: {e}")
            return False

        # Build tests
        print("🔨 Building tests...")
        try:
            subprocess.run(["cmake", "--build", "."], cwd=self.output_dir, check=True)
        except subprocess.CalledProcessError as e:
            print(f"❌ Build failed: {e}")
            return False

        # Run tests
        print("🧪 Running tests...")
        try:
            ctest_cmd = ["ctest", "--output-on-failure"]
            if self.ctest_regex:
                ctest_cmd.extend(["-R", self.ctest_regex])

            env = os.environ.copy()
            # If v2 section approvals exist, restrict execution to approved sections in-scope.
            gtest_filter = _compute_gtest_filter_for_scope(self.repo_path, ctest_regex=self.ctest_regex)
            if gtest_filter:
                env["GTEST_FILTER"] = gtest_filter

            result = subprocess.run(ctest_cmd, cwd=self.output_dir, capture_output=True, text=True, env=env)
            combined_output = (result.stdout or "") + (result.stderr or "")
            no_tests_found = "No tests were found" in combined_output

            # Best-effort: extract how many tests GoogleTest reported running.
            ran_count: int | None = None
            try:
                m = re.search(r"\[=+\]\s*Running\s+(\d+)\s+tests?\s+from\s+\d+\s+test\s+suites?\.", combined_output)
                if m:
                    ran_count = int(m.group(1))
            except Exception:
                ran_count = None

            test_results = [{
                "passed": (result.returncode == 0) and (not no_tests_found),
                "output": combined_output,
            }]
            # Emit per-test-case report for GoogleTest executables (function-scenario granularity).
            try:
                self._write_gtest_case_reports(result)
            except Exception:
                pass

            # CTest data is now included in the interlocking_test_report.txt, no separate file needed
        except subprocess.CalledProcessError as e:
            print(f"❌ Tests failed: {e}")
            test_results = []

        if not test_results:
            print("❌ Tests did not run")
            return False

        if no_tests_found or ran_count == 0:
            print("❌ No tests were executed")
            return False

        if not test_results[0]["passed"]:
            print("❌ Tests failed")
            return False

        print("✅ All tests passed!")
        print(f"📄 Test Output:\n{test_results[0]['output']}")

        return True

    def _write_gtest_case_reports(self, ctest_result=None):
        """Run each built gtest executable with XML output and summarize per test case."""
        import xml.etree.ElementTree as ET
        import datetime
        import csv
        tests_bin_dir = self.output_dir / "tests"
        if not tests_bin_dir.exists():
            return

        # Keep reports consistent with any approvals-based filtering.
        env = os.environ.copy()
        if self.ctest_regex:
            gtest_filter = _compute_gtest_filter_for_scope(self.repo_path, ctest_regex=self.ctest_regex)
            if gtest_filter:
                env["GTEST_FILTER"] = gtest_filter

        exes = [p for p in tests_bin_dir.iterdir() if p.is_file() and p.suffix.lower() == ".exe"]
        if not exes:
            # Non-Windows builds may have no suffix.
            exes = [p for p in tests_bin_dir.iterdir() if p.is_file() and p.suffix == "" and not p.name.endswith(".cmake")]
        if not exes:
            return

        report_root = self._report_dir()
        self._clean_report_root(report_root)
        run_started_utc = datetime.datetime.now(datetime.timezone.utc)
        run_summaries: list[dict] = []


        combined_ctest_output = ""
        ctest_no_tests_found = False
        ctest_return_code = None
        if ctest_result is not None:
            combined_ctest_output = (ctest_result.stdout or "") + (ctest_result.stderr or "")
            ctest_no_tests_found = "No tests were found" in combined_ctest_output
            ctest_return_code = ctest_result.returncode

        def _ctest_failure_reason() -> str:
            if ctest_result is None:
                return "CTest data not available"
            if ctest_no_tests_found:
                return "CTest discovered 0 tests (\"No tests were found\")"
            if ctest_return_code is not None and ctest_return_code != 0:
                return f"CTest returned non-zero exit code ({ctest_return_code})"
            return ""

        def _ctest_output_excerpt(max_lines: int = 60) -> str:
            if not combined_ctest_output:
                return ""
            lines = combined_ctest_output.replace("\r\n", "\n").split("\n")
            lines = [ln for ln in lines if ln.strip()]
            excerpt = "\n".join(lines[:max_lines])
            if len(lines) > max_lines:
                excerpt += "\n... (truncated)"
            return excerpt

        def _per_exe_report_dir(exe_name: str) -> Path:
            """Place reports under tests/test_reports/src/<parents>/<test_file_stem>/.

            Desired layout:
              tests/test_reports/src/<parents>/<test_file_stem>/
                - gtest.xml
                - <test_file_stem>_test_report.txt
            """

            # Map executable name back to its source test path.
            # CMake may name exes like: test_Foo__src__bar (stem + '__' + dir with '/' -> '__').
            stem = exe_name.split("__", 1)[0]
            dir_safe = exe_name.split("__", 1)[1] if "__" in exe_name else ""

            test_src: Path | None = None
            if dir_safe:
                rel_dir = Path(*[p for p in dir_safe.split("__") if p])
                for ext in (".cpp", ".cc", ".cxx", ".c++", ".c"):
                    candidate = self.tests_dir / rel_dir / f"{stem}{ext}"
                    if candidate.exists() and candidate.is_file():
                        test_src = candidate
                        break

            if test_src is None:
                for ext in (".cpp", ".cc", ".cxx", ".c++", ".c"):
                    hits = [p for p in self.tests_dir.rglob(f"{stem}{ext}") if p.is_file()]
                    if hits:
                        test_src = hits[0]
                        break
            if test_src is None:
                report_dir = report_root
            else:
                try:
                    rel = test_src.relative_to(self.tests_dir)
                    # Use the file stem as the final folder segment (avoid .cpp folders).
                    report_dir = report_root / rel.parent / test_src.stem
                except Exception:
                    report_dir = report_root

            report_dir.mkdir(parents=True, exist_ok=True)
            return report_dir

        def _report_filename_for_exe(exe_name: str) -> str:
            # Prefer the test file stem if we can resolve it; otherwise fall back to exe stem.
            stem = exe_name.split("__", 1)[0]
            dir_safe = exe_name.split("__", 1)[1] if "__" in exe_name else ""

            test_src: Path | None = None
            if dir_safe:
                rel_dir = Path(*[p for p in dir_safe.split("__") if p])
                for ext in (".cpp", ".cc", ".cxx", ".c++", ".c"):
                    candidate = self.tests_dir / rel_dir / f"{stem}{ext}"
                    if candidate.exists() and candidate.is_file():
                        test_src = candidate
                        break

            if test_src is None:
                for ext in (".cpp", ".cc", ".cxx", ".c++", ".c"):
                    hits = [p for p in self.tests_dir.rglob(f"{stem}{ext}") if p.is_file()]
                    if hits:
                        test_src = hits[0]
                        break

            report_stem = test_src.stem if test_src is not None else stem
            return f"{report_stem}_test_report.txt"

        def _report_stem_for_exe(exe_name: str) -> str:
            # The test file stem (preferred), otherwise the exe stem.
            stem = exe_name.split("__", 1)[0]
            dir_safe = exe_name.split("__", 1)[1] if "__" in exe_name else ""

            test_src: Path | None = None
            if dir_safe:
                rel_dir = Path(*[p for p in dir_safe.split("__") if p])
                for ext in (".cpp", ".cc", ".cxx", ".c++", ".c"):
                    candidate = self.tests_dir / rel_dir / f"{stem}{ext}"
                    if candidate.exists() and candidate.is_file():
                        test_src = candidate
                        break

            if test_src is None:
                for ext in (".cpp", ".cc", ".cxx", ".c++", ".c"):
                    hits = [p for p in self.tests_dir.rglob(f"{stem}{ext}") if p.is_file()]
                    if hits:
                        test_src = hits[0]
                        break

            return test_src.stem if test_src is not None else stem

        for exe in exes:
            per_dir = _per_exe_report_dir(exe.name)
            report_stem = _report_stem_for_exe(exe.name)
            xml_path = per_dir / "gtest.xml"
            try:
                run = subprocess.run(
                    [str(exe), f"--gtest_output=xml:{xml_path}"],
                    cwd=tests_bin_dir,
                    capture_output=True,
                    text=True,
                    timeout=60,
                    env=env,
                )
            except Exception:
                continue

            # Even if the executable fails, try to parse whatever it produced.
            if not xml_path.exists():
                continue

            try:
                root = ET.parse(xml_path).getroot()
            except Exception:
                continue

            cases: list[dict] = []

            # GoogleTest XML structure: <testsuites><testsuite name=...><testcase name=...><failure .../></testcase>...
            for suite in root.findall(".//testsuite"):
                suite_name = suite.attrib.get("name", "")
                for case in suite.findall("testcase"):
                    case_name = case.attrib.get("name", "")
                    failures = case.findall("failure")
                    status = "FAILED" if failures else "PASSED"
                    failure_msg = ""
                    failure_full = ""
                    if failures:
                        f0 = failures[0]
                        failure_msg = (f0.attrib.get("message") or "").strip()
                        failure_full = ((f0.text or "") or "").strip()
                        if not failure_msg and failure_full:
                            failure_msg = failure_full.splitlines()[0]
                    cases.append({
                        "suite": suite_name,
                        "case": case_name,
                        "status": status,
                        "failure": failure_msg,
                        "failure_full": failure_full,
                    })

            if not cases:
                continue

            # Write a readable summary per executable.
            summary_path = per_dir / _report_filename_for_exe(exe.name)
            with open(summary_path, "w", encoding="utf-8") as f:
                f.write("=" * 60 + "\n")
                f.write(f"GTEST CASE REPORT\n")
                f.write("=" * 60 + "\n\n")
                f.write(f"Run (UTC): {run_started_utc.isoformat()}\n")
                f.write(f"Group: {self.report_group}\n")
                f.write(f"Executable: {exe.name}\n\n")
                passed = sum(1 for c in cases if c["status"] == "PASSED")
                failed = sum(1 for c in cases if c["status"] == "FAILED")
                f.write(f"Total cases: {len(cases)}\n")
                f.write(f"Passed: {passed}\n")
                f.write(f"Failed: {failed}\n\n")

                # Failed-first listing (most actionable at top).
                ordered = sorted(
                    cases,
                    key=lambda c: (0 if c.get("status") == "FAILED" else 1, c.get("suite", ""), c.get("case", "")),
                )
                for c in ordered:
                    line = f"{c['status']}: {c['suite']}.{c['case']}"
                    if c.get("failure"):
                        line += f"  |  {c['failure']}"
                    f.write(line + "\n")
                
                # Add CTest summary
                f.write("\n\n" + "=" * 60 + "\n")
                f.write("CTEST SUMMARY\n")
                f.write("=" * 60 + "\n\n")
                if ctest_result:
                    f.write(f"Return code: {ctest_result.returncode}\n\n")
                    if ctest_no_tests_found:
                        f.write("STATUS: FAILED (no tests were discovered by CTest)\n\n")
                        if self.ctest_regex:
                            f.write(f"Note: You passed CTEST -R: {self.ctest_regex}\n")
                            f.write("If the regex doesn't match any CTest names, zero tests run.\n\n")
                    else:
                        f.write("STATUS: COMPLETED\n\n")
                    f.write("--- STDOUT/STDERR ---\n")
                    f.write(ctest_result.stdout or "")
                    if ctest_result.stderr:
                        f.write("\n--- STDERR ---\n")
                        f.write(ctest_result.stderr)
                else:
                    f.write("(CTest data not available)\n")

            # Print the report path (relative to repo) for quick discovery.
            try:
                rel_print = summary_path.relative_to(self.repo_path).as_posix()
            except Exception:
                rel_print = str(summary_path)
            print(f"   📄 Wrote report: {rel_print}")

            # Write per-file SUMMARY and Excel/CSV results next to the report.
            per_summary = per_dir / f"{report_stem}_SUMMARY.txt"
            with open(per_summary, "w", encoding="utf-8") as f:
                f.write("=" * 60 + "\n")
                f.write("AI TEST REPORT – SUMMARY\n")
                f.write("=" * 60 + "\n\n")
                f.write(f"Run (UTC): {run_started_utc.isoformat()}\n")
                f.write(f"Group: {self.report_group}\n")
                if self.ctest_regex:
                    f.write(f"CTEST -R: {self.ctest_regex}\n")
                if env.get("GTEST_FILTER"):
                    f.write(f"GTEST_FILTER: {env.get('GTEST_FILTER')}\n")
                if ctest_return_code is not None:
                    f.write(f"CTEST return code: {ctest_return_code}\n")
                if ctest_no_tests_found:
                    f.write("CTEST status: FAILED (no tests were discovered by CTest)\n")
                reason = _ctest_failure_reason()
                if reason:
                    f.write(f"RUN FAILURE REASON: {reason}\n")
                    if ctest_no_tests_found and self.ctest_regex:
                        f.write("Likely cause: -R regex didn't match any CTest test names, or tests were not registered via add_test().\n")
                    elif ctest_return_code is not None and ctest_return_code != 0 and not ctest_no_tests_found:
                        f.write("Likely cause: one or more test cases failed (see FAILED list in the report).\n")
                excerpt = _ctest_output_excerpt()
                if excerpt:
                    f.write("\nCTEST OUTPUT (excerpt)\n")
                    f.write("-" * 20 + "\n")
                    f.write(excerpt + "\n")
                f.write("\n")
                f.write(f"Executable: {exe.name}\n")
                f.write(f"Total cases: {len(cases)}\n")
                f.write(f"Passed: {passed}\n")
                f.write(f"Failed: {failed}\n")

            try:
                rel_print = per_summary.relative_to(self.repo_path).as_posix()
            except Exception:
                rel_print = str(per_summary)
            print(f"   📄 Wrote report: {rel_print}")

            per_case_rows: list[dict] = []
            for c in cases:
                # Keep full failure text readable in Excel by truncating.
                failure_full = (c.get("failure_full") or "").replace("\r\n", "\n").strip()
                if failure_full:
                    flines = [ln for ln in failure_full.split("\n") if ln.strip()]
                    failure_excerpt = "\n".join(flines[:10])
                    if len(flines) > 10:
                        failure_excerpt += "\n... (truncated)"
                else:
                    failure_excerpt = ""

                per_case_rows.append(
                    {
                        "run_utc": run_started_utc.isoformat(),
                        "ctest_regex": self.ctest_regex or "",
                        "gtest_filter": env.get("GTEST_FILTER", ""),
                        "ctest_return_code": "" if ctest_return_code is None else str(ctest_return_code),
                        "ctest_no_tests_found": str(bool(ctest_no_tests_found)),
                        "run_failure_reason": _ctest_failure_reason(),
                        "ctest_output_excerpt": _ctest_output_excerpt(40),
                        "executable": exe.name,
                        "suite": c.get("suite", ""),
                        "case": c.get("case", ""),
                        "status": c.get("status", ""),
                        "failure": c.get("failure", ""),
                        "failure_details": failure_excerpt,
                    }
                )

            if per_case_rows:
                fieldnames = list(per_case_rows[0].keys())
                # Prefer XLSX for testers; fall back to CSV if openpyxl isn't available.
                wrote_xlsx = False
                try:
                    import openpyxl  # type: ignore
                    from openpyxl.styles import Font  # type: ignore

                    per_results_xlsx = per_dir / f"{report_stem}_RESULTS.xlsx"
                    wb = openpyxl.Workbook()

                    ws = wb.active
                    ws.title = "Results"
                    ws.append(fieldnames)
                    for cell in ws[1]:
                        cell.font = Font(bold=True)
                    for row in per_case_rows:
                        ws.append([row.get(k, "") for k in fieldnames])
                    ws.freeze_panes = "A2"

                    ws2 = wb.create_sheet("Summary")
                    ws2.append(["Key", "Value"])
                    for cell in ws2[1]:
                        cell.font = Font(bold=True)
                    ws2.append(["run_utc", run_started_utc.isoformat()])
                    ws2.append(["group", self.report_group])
                    ws2.append(["ctest_regex", self.ctest_regex or ""])
                    ws2.append(["gtest_filter", env.get("GTEST_FILTER", "")])
                    ws2.append(["ctest_return_code", "" if ctest_return_code is None else str(ctest_return_code)])
                    ws2.append(["ctest_no_tests_found", str(bool(ctest_no_tests_found))])
                    ws2.append(["run_failure_reason", _ctest_failure_reason()])
                    ws2.append(["ctest_output_excerpt", _ctest_output_excerpt(80)])

                    wb.save(per_results_xlsx)
                    wrote_xlsx = True
                    try:
                        rel_print = per_results_xlsx.relative_to(self.repo_path).as_posix()
                    except Exception:
                        rel_print = str(per_results_xlsx)
                    print(f"   📄 Wrote report: {rel_print}")
                except Exception:
                    wrote_xlsx = False

                if not wrote_xlsx:
                    per_results_csv = per_dir / f"{report_stem}_RESULTS.csv"
                    with open(per_results_csv, "w", newline="", encoding="utf-8") as f:
                        writer = csv.DictWriter(f, fieldnames=fieldnames)
                        writer.writeheader()
                        writer.writerows(per_case_rows)
                    try:
                        rel_print = per_results_csv.relative_to(self.repo_path).as_posix()
                    except Exception:
                        rel_print = str(per_results_csv)
                    print(f"   📄 Wrote report: {rel_print}")

            run_summaries.append(
                {
                    "exe": exe.name,
                    "passed": passed,
                    "failed": failed,
                    "total": len(cases),
                    "report": str(summary_path.relative_to(report_root)).replace("\\", "/")
                    if summary_path.is_relative_to(report_root)
                    else str(summary_path),
                }
            )

def main():
    """Main entry point for the AI Test Runner."""
    parser = argparse.ArgumentParser(
        description="AI Test Runner - Compiles, executes, and provides coverage for AI-generated C/C++ unit tests"
    )
    parser.add_argument(
        "repo_path",
        help="Path to the repository containing tests"
    )
    parser.add_argument(
        "--output-dir",
        default="build",
        help="Output directory name under <repo>/tests/ (default: build -> tests/build)"
    )
    parser.add_argument(
        "--ctest-regex",
        default=None,
        help="Optional: pass a regex to ctest via -R to run only matching tests",
    )
    parser.add_argument(
        "--language",
        choices=["c", "cpp", "auto"],
        default="auto",
        help="Programming language (default: auto-detect)"
    )
    parser.add_argument(
        "--version",
        action="version",
        version="%(prog)s 1.0.0"
    )

    parser.add_argument(
        "--safety-level",
        choices=list(SafetyPolicy.allowed_levels()),
        default="QM",
        help=(
            "Configures which analyses, test types, and review gates are required so generated tests align with SIL expectations "
            "without claiming certification."
        ),
    )
    parser.add_argument("--policy-file", default=None)
    parser.add_argument(
        "--disable-mcdc",
        action="store_true",
        help="Advanced override: disable MC/DC expectations even if the selected safety level would enable them.",
    )

    args = parser.parse_args()

    repo_root = Path(args.repo_path).resolve()
    try:
        policy = SafetyPolicy.load(
            safety_level=args.safety_level,
            repo_root=repo_root,
            policy_file=args.policy_file,
            disable_mcdc=bool(args.disable_mcdc),
        )
    except Exception:
        # If policy loading fails, default to the safest stance: require approvals.
        policy = SafetyPolicy(
            safety_level=str(args.safety_level or "QM"),
            approval_required=True,
            coverage_target={},
            mcdc_analysis=False,
            mcdc_generation=False,
        )

    # Enforce approvals whenever v2 registry exists (source of truth).
    # This prevents starting a build when pending active sections exist.
    _enforce_section_review_gate(repo_root, ctest_regex=args.ctest_regex)

    # Create and run the test runner
    runner = AITestRunner(args.repo_path, args.output_dir, args.language)
    runner.ctest_regex = args.ctest_regex
    success = runner.run()

    # Best-effort: update safety summary.
    try:
        save_safety_summary(
            repo_root,
            {
                "safety_level": policy.safety_level,
                "human_approvals_complete": bool(policy.approval_required),
                "coverage_status": {
                    "statement": "NOT_RUN",
                    "branch": "NOT_RUN",
                    "mcdc": "NOT_RUN" if policy.mcdc_expected() else "N/A",
                },
            },
        )
    except Exception:
        pass

    # Exit with appropriate code
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
